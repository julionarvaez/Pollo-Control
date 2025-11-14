"""
Sistema de Control de Pollos de Engorde - Backend
API REST con Flask
"""
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, timezone, date
from functools import wraps
import jwt
import os
import warnings
import io
import importlib

# Intentar importar qrcode (opcional)
try:
    import qrcode
    QRCODE_AVAILABLE = True
except ImportError:
    QRCODE_AVAILABLE = False
    print("⚠️  Librería qrcode no disponible. Instalar con: pip install qrcode[pil]")

# Suprimir warnings de deprecación de datetime.utcnow()
warnings.filterwarnings("ignore", category=DeprecationWarning, message="datetime.datetime.utcnow*")

# Inicialización: servir también el frontend estático en producción
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.abspath(os.path.join(BASE_DIR, '..', 'frontend'))

app = Flask(
    __name__,
    static_folder=FRONTEND_DIR,
    static_url_path=''
)
# Configuración CORS simplificada y más permisiva para entorno local.
# Nota: supports_credentials implica que no podemos usar '*' como origen; Flask-CORS se encarga de reflejar el origen permitido.
CORS(app,
    resources={r"/api/*": {"origins": ["http://127.0.0.1:5500", "http://localhost:5500", "http://127.0.0.1:8000", "http://localhost:8000", "http://localhost", "http://127.0.0.1", "null"]}},
     supports_credentials=True,
     expose_headers=["Content-Disposition"],
     allow_headers=["Content-Type", "Authorization", "X-Requested-With"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
     max_age=600)

# Configuración
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'clave-super-secreta-2025')

# Normalizar DATABASE_URL (Render usa postgres://)
db_url = os.environ.get('DATABASE_URL', 'sqlite:///pollo_control.db')
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql+psycopg2://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ============= MODELOS =============

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    nombre_completo = db.Column(db.String(100))
    email = db.Column(db.String(100))
    rol = db.Column(db.String(20), default='admin')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Lote(db.Model):
    __tablename__ = 'lotes'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    fecha_inicio = db.Column(db.Date, nullable=False)
    fecha_fin = db.Column(db.Date)
    cantidad_inicial = db.Column(db.Integer, nullable=False)
    cantidad_actual = db.Column(db.Integer)
    genetica = db.Column(db.String(50))
    proveedor = db.Column(db.String(100))
    peso_inicial = db.Column(db.Float)
    galpon = db.Column(db.String(50))
    dias_ciclo = db.Column(db.Integer, default=42)  # Duración típica del ciclo en días
    estado = db.Column(db.String(20), default='activo')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    registros = db.relationship('RegistroDiario', backref='lote', lazy=True, cascade='all, delete-orphan')
    costos = db.relationship('Costo', backref='lote', lazy=True, cascade='all, delete-orphan')
    ingresos = db.relationship('Ingreso', backref='lote', lazy=True, cascade='all, delete-orphan')
    sanidad = db.relationship('Sanidad', backref='lote', lazy=True, cascade='all, delete-orphan')

class RegistroDiario(db.Model):
    __tablename__ = 'registros_diarios'
    id = db.Column(db.Integer, primary_key=True)
    lote_id = db.Column(db.Integer, db.ForeignKey('lotes.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    alimento_kg = db.Column(db.Float)
    agua_litros = db.Column(db.Float)
    mortalidad = db.Column(db.Integer, default=0)
    causa_mortalidad = db.Column(db.Text)
    peso_promedio = db.Column(db.Float)
    temperatura_promedio = db.Column(db.Float)
    humedad = db.Column(db.Float)  # Humedad relativa (%) del galpón
    observaciones = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    __table_args__ = (db.UniqueConstraint('lote_id', 'fecha', name='_lote_fecha_uc'),)

class Costo(db.Model):
    __tablename__ = 'costos'
    id = db.Column(db.Integer, primary_key=True)
    lote_id = db.Column(db.Integer, db.ForeignKey('lotes.id'), nullable=False)
    categoria = db.Column(db.String(50), nullable=False)
    concepto = db.Column(db.String(200), nullable=False)
    monto = db.Column(db.Float, nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    observaciones = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Ingreso(db.Model):
    __tablename__ = 'ingresos'
    id = db.Column(db.Integer, primary_key=True)
    lote_id = db.Column(db.Integer, db.ForeignKey('lotes.id'), nullable=False)
    cantidad_vendida = db.Column(db.Integer, nullable=False)
    peso_promedio = db.Column(db.Float, nullable=False)
    precio_por_kg = db.Column(db.Float, nullable=False)
    total = db.Column(db.Float, nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    cliente = db.Column(db.String(100))
    observaciones = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Sanidad(db.Model):
    __tablename__ = 'sanidad'
    id = db.Column(db.Integer, primary_key=True)
    lote_id = db.Column(db.Integer, db.ForeignKey('lotes.id'), nullable=False)
    tipo = db.Column(db.String(50), nullable=False)
    producto = db.Column(db.String(200), nullable=False)
    dosis = db.Column(db.String(100))
    fecha = db.Column(db.Date, nullable=False)
    edad_dias = db.Column(db.Integer)
    via_administracion = db.Column(db.String(50))  # Oral, Agua, Inyectable, Spray, Alimento
    retiro_dias = db.Column(db.Integer)            # Días de retiro antes de sacrificio
    enfermedad_id = db.Column(db.Integer, db.ForeignKey('enfermedades.id'))
    enfermedad = db.relationship('Enfermedad')
    observaciones = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Configuracion(db.Model):
    __tablename__ = 'configuracion'
    id = db.Column(db.Integer, primary_key=True)
    temp_min = db.Column(db.Float, default=22.0)
    temp_max = db.Column(db.Float, default=32.0)
    humedad_min = db.Column(db.Float, default=50.0)
    humedad_max = db.Column(db.Float, default=70.0)
    consumo_min_g_dia = db.Column(db.Float, default=60.0)  # g/ave/día
    mortalidad_diaria_max_pct = db.Column(db.Float, default=1.0)  # % diaria
    peso_tolerancia_pct = db.Column(db.Float, default=5.0)  # Banda ±% sobre peso objetivo
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Enfermedad(db.Model):
    __tablename__ = 'enfermedades'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(120), unique=True, nullable=False)
    sintomas = db.Column(db.Text)
    prevencion = db.Column(db.Text)
    tratamiento = db.Column(db.Text)
    medicamentos = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ============= FUNCIONES DE INICIALIZACIÓN =============

def ensure_database_schema():
    """Garantiza que existan columnas añadidas por migraciones simples."""
    try:
        with app.app_context():
            engine = db.engine
            inspector = inspect(engine)
            tables = set(inspector.get_table_names())

            if 'sanidad' in tables:
                sanidad_cols = {col['name'] for col in inspector.get_columns('sanidad')}
                missing_statements = []

                if 'via_administracion' not in sanidad_cols:
                    missing_statements.append(('via_administracion', "ALTER TABLE sanidad ADD COLUMN via_administracion TEXT"))
                if 'retiro_dias' not in sanidad_cols:
                    missing_statements.append(('retiro_dias', "ALTER TABLE sanidad ADD COLUMN retiro_dias INTEGER"))
                if 'enfermedad_id' not in sanidad_cols:
                    missing_statements.append(('enfermedad_id', "ALTER TABLE sanidad ADD COLUMN enfermedad_id INTEGER"))

                if missing_statements:
                    with engine.begin() as conn:
                        for _, stmt in missing_statements:
                            conn.execute(text(stmt))
                    added = ', '.join(name for name, _ in missing_statements)
                    print(f'✅ Columnas faltantes agregadas a sanidad: {added}')
    except Exception as exc:
        # Registrar el problema pero no bloquear el arranque del backend
        print(f"⚠️  No se pudieron aplicar migraciones automáticas: {exc}")


def ensure_app_dirs():
    try:
        # Directorio de datos (persistente si se usa DATA_DIR montado en Render)
        data_dir = os.environ.get('DATA_DIR', os.path.join(BASE_DIR, 'instance'))
        paths = [
            os.path.join(data_dir, 'uploads'),
            os.path.join(data_dir, 'exports'),
            os.path.join(data_dir, 'backups'),
        ]
        for p in paths:
            os.makedirs(p, exist_ok=True)
        # Apuntar config a estas rutas
        app.config.setdefault('UPLOAD_FOLDER', paths[0])
        app.config.setdefault('EXPORT_FOLDER', paths[1])
        app.config.setdefault('BACKUP_FOLDER', paths[2])
    except Exception as exc:
        print(f"⚠️  No se pudieron crear directorios de trabajo: {exc}")

def bootstrap_database():
    try:
        with app.app_context():
            db.create_all()
            if not Usuario.query.filter_by(username='admin').first():
                admin = Usuario(
                    username='admin',
                    password_hash=generate_password_hash('admin123'),
                    nombre_completo='Administrador',
                    email='admin@granja.com',
                    rol='admin'
                )
                db.session.add(admin)
                db.session.commit()
                print("✅ Usuario administrador creado: admin / admin123")
    except Exception as exc:
        print(f"⚠️  Bootstrap de base de datos falló: {exc}")

# Inicializar todo
ensure_database_schema()
ensure_app_dirs()
bootstrap_database()

@app.route('/api/ping', methods=['GET'])
def ping():
    """Ruta simple para probar disponibilidad y CORS sin autenticación."""
    return jsonify({'pong': True, 'timestamp': datetime.utcnow().isoformat()})

@app.route('/api/<path:any_path>', methods=['OPTIONS'])
def generic_options(any_path):
    """Maneja preflight OPTIONS devolviendo las cabeceras necesarias."""
    resp = jsonify({'ok': True})
    return resp, 200

@app.route('/favicon.ico')
def favicon():
    # Evita el 404 del navegador cuando solicita favicon desde el servidor estático
    return ('', 204)

# ============= DECORADORES =============

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'mensaje': 'Token no proporcionado'}), 401
        
        try:
            if token.startswith('Bearer '):
                token = token[7:]
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = Usuario.query.get(data['user_id'])
            if not current_user:
                return jsonify({'mensaje': 'Usuario no encontrado'}), 401
        except jwt.ExpiredSignatureError:
            return jsonify({'mensaje': 'Token expirado'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'mensaje': 'Token inválido'}), 401
        except Exception as e:
            return jsonify({'mensaje': f'Error de autenticación: {str(e)}'}), 401
        
        return f(current_user, *args, **kwargs)
    return decorated

# ============= SERVICIOS DE CÁLCULO =============

def calcular_estadisticas(lote):
    """Calcula todas las estadísticas del lote"""
    try:
        dias_transcurridos = (datetime.now().date() - lote.fecha_inicio).days
        
        # Totales básicos
        total_alimento = sum(r.alimento_kg or 0 for r in lote.registros)
        total_mortalidad = sum(r.mortalidad or 0 for r in lote.registros)
        total_agua = sum(r.agua_litros or 0 for r in lote.registros)
        
        # Peso
        registros_con_peso = sorted([r for r in lote.registros if r.peso_promedio], key=lambda x: x.fecha)
        peso_inicial = lote.peso_inicial if lote.peso_inicial else 40
        peso_actual = registros_con_peso[-1].peso_promedio if registros_con_peso else peso_inicial
        
        # Cantidad actual
        cantidad_actual = max(0, lote.cantidad_inicial - total_mortalidad)
        
        # Ganancia de peso
        ganancia_peso = max(0, peso_actual - peso_inicial)
        ganancia_peso_total = ganancia_peso * cantidad_actual / 1000  # en kg
        
        # FCR (Feed Conversion Ratio)
        fcr = total_alimento / ganancia_peso_total if ganancia_peso_total > 0 else 0
        
        # ADG (Average Daily Gain)
        adg = ganancia_peso / dias_transcurridos if dias_transcurridos > 0 else 0
        
        # Mortalidad %
        mortalidad_porcentaje = (total_mortalidad / lote.cantidad_inicial * 100) if lote.cantidad_inicial > 0 else 0
        
        # Economía
        total_costos = sum(c.monto for c in lote.costos)
        total_ingresos = sum(i.total for i in lote.ingresos)
        ganancia = total_ingresos - total_costos
        
        # Costo por kg
        kg_producidos = ganancia_peso_total
        costo_por_kg = total_costos / kg_producidos if kg_producidos > 0 else 0
        
        # Rentabilidad
        rentabilidad = (ganancia / total_costos * 100) if total_costos > 0 else 0
        
        # Consumo promedio
        consumo_promedio_diario = total_alimento / dias_transcurridos if dias_transcurridos > 0 else 0
        consumo_por_ave = total_alimento * 1000 / cantidad_actual if cantidad_actual > 0 else 0
        
        # Relación agua/alimento
        agua_alimento = total_agua / total_alimento if total_alimento > 0 else 0
        
        return {
            'dias_transcurridos': dias_transcurridos,
            'cantidad_actual': cantidad_actual,
            'peso_actual': round(peso_actual, 2),
            'total_alimento': round(total_alimento, 2),
            'total_agua': round(total_agua, 2),
            'total_mortalidad': total_mortalidad,
            'mortalidad_porcentaje': round(mortalidad_porcentaje, 2),
            'fcr': round(fcr, 2),
            'adg': round(adg, 2),
            'kg_producidos': round(kg_producidos, 2),
            'total_costos': round(total_costos, 2),
            'total_ingresos': round(total_ingresos, 2),
            'ganancia': round(ganancia, 2),
            'costo_por_kg': round(costo_por_kg, 2),
            'costo_por_pollo': round(total_costos / lote.cantidad_inicial, 2) if lote.cantidad_inicial > 0 else 0,
            'rentabilidad': round(rentabilidad, 2),
            'consumo_promedio_diario': round(consumo_promedio_diario, 2),
            'consumo_por_ave': round(consumo_por_ave, 2),
            'agua_alimento': round(agua_alimento, 2)
        }
    except Exception as e:
        print(f"Error en calcular_estadisticas para lote {lote.id}: {str(e)}")
        import traceback
        traceback.print_exc()
        # Retornar valores por defecto en caso de error
        return {
            'dias_transcurridos': 0,
            'cantidad_actual': lote.cantidad_inicial,
            'peso_actual': lote.peso_inicial or 40,
            'total_alimento': 0,
            'total_agua': 0,
            'total_mortalidad': 0,
            'mortalidad_porcentaje': 0,
            'fcr': 0,
            'adg': 0,
            'kg_producidos': 0,
            'total_costos': 0,
            'total_ingresos': 0,
            'ganancia': 0,
            'costo_por_kg': 0,
            'costo_por_pollo': 0,
            'rentabilidad': 0,
            'consumo_promedio_diario': 0,
            'consumo_por_ave': 0,
            'agua_alimento': 0
        }

def get_configuracion_valores():
    """Obtiene la configuración de umbrales; crea una con valores por defecto si no existe."""
    cfg = Configuracion.query.first()
    if not cfg:
        cfg = Configuracion()
        db.session.add(cfg)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
    return {
        'temp_min': cfg.temp_min if cfg else 22.0,
        'temp_max': cfg.temp_max if cfg else 32.0,
        'humedad_min': cfg.humedad_min if cfg else 50.0,
        'humedad_max': cfg.humedad_max if cfg else 70.0,
        'consumo_min_g_dia': cfg.consumo_min_g_dia if cfg else 60.0,
        'mortalidad_diaria_max_pct': cfg.mortalidad_diaria_max_pct if cfg else 1.0,
        'peso_tolerancia_pct': cfg.peso_tolerancia_pct if cfg else 5.0,
    }

# ============= RUTAS - AUTENTICACIÓN =============

@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        
        if not data or not data.get('username') or not data.get('password'):
            return jsonify({'mensaje': 'Faltan credenciales'}), 400
        
        usuario = Usuario.query.filter_by(username=data.get('username')).first()
        
        if not usuario or not check_password_hash(usuario.password_hash, data.get('password')):
            return jsonify({'mensaje': 'Credenciales inválidas'}), 401
        
        token = jwt.encode({
            'user_id': usuario.id,
            'exp': datetime.utcnow() + timedelta(days=30)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        
        return jsonify({
            'token': token,
            'usuario': {
                'id': usuario.id,
                'username': usuario.username,
                'nombre_completo': usuario.nombre_completo,
                'email': usuario.email,
                'rol': usuario.rol
            }
        })
    except Exception as e:
        return jsonify({'mensaje': f'Error en login: {str(e)}'}), 500

@app.route('/api/auth/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        
        if not data or not data.get('username') or not data.get('password'):
            return jsonify({'mensaje': 'Faltan datos requeridos'}), 400
        
        if Usuario.query.filter_by(username=data.get('username')).first():
            return jsonify({'mensaje': 'Usuario ya existe'}), 400
        
        usuario = Usuario(
            username=data.get('username'),
            password_hash=generate_password_hash(data.get('password')),
            nombre_completo=data.get('nombre_completo'),
            email=data.get('email')
        )
        
        db.session.add(usuario)
        db.session.commit()
        
        return jsonify({'mensaje': 'Usuario creado exitosamente'}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al crear usuario: {str(e)}'}), 500

# ============= RUTAS - LOTES =============

@app.route('/api/lotes', methods=['GET'])
@token_required
def get_lotes(current_user):
    try:
        lotes = Lote.query.order_by(Lote.fecha_inicio.desc()).all()
        return jsonify([{
            'id': l.id,
            'nombre': l.nombre,
            'fecha_inicio': l.fecha_inicio.isoformat(),
            'fecha_fin': l.fecha_fin.isoformat() if l.fecha_fin else None,
            'cantidad_inicial': l.cantidad_inicial,
            'cantidad_actual': l.cantidad_actual or l.cantidad_inicial,
            'estado': l.estado,
            'genetica': l.genetica,
            'proveedor': l.proveedor,
            'galpon': l.galpon,
            'dias_ciclo': l.dias_ciclo or 42,
            'dias_transcurridos': (datetime.now().date() - l.fecha_inicio).days,
            'dias_restantes': (l.dias_ciclo or 42) - (datetime.now().date() - l.fecha_inicio).days,
            'fecha_sacrificio': (l.fecha_inicio + timedelta(days=l.dias_ciclo or 42)).isoformat()
        } for l in lotes])
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener lotes: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>', methods=['GET'])
@token_required
def get_lote(current_user, id):
    try:
        lote = Lote.query.get_or_404(id)
        return jsonify({
            'id': lote.id,
            'nombre': lote.nombre,
            'fecha_inicio': lote.fecha_inicio.isoformat(),
            'fecha_fin': lote.fecha_fin.isoformat() if lote.fecha_fin else None,
            'cantidad_inicial': lote.cantidad_inicial,
            'cantidad_actual': lote.cantidad_actual or lote.cantidad_inicial,
            'genetica': lote.genetica,
            'proveedor': lote.proveedor,
            'peso_inicial': lote.peso_inicial,
            'galpon': lote.galpon,
            'estado': lote.estado,
            'dias_ciclo': lote.dias_ciclo or 42,
            'dias_transcurridos': (datetime.now().date() - lote.fecha_inicio).days,
            'dias_restantes': (lote.dias_ciclo or 42) - (datetime.now().date() - lote.fecha_inicio).days,
            'fecha_sacrificio': (lote.fecha_inicio + timedelta(days=lote.dias_ciclo or 42)).isoformat()
        })
    except Exception as e:
        print(f"Error al obtener lote {id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'mensaje': f'Error al obtener lote: {str(e)}'}), 500

@app.route('/api/lotes', methods=['POST'])
@token_required
def crear_lote(current_user):
    try:
        data = request.get_json()
        
        if not data or not data.get('nombre') or not data.get('fecha_inicio') or not data.get('cantidad_inicial'):
            return jsonify({'mensaje': 'Faltan datos requeridos'}), 400
        
        lote = Lote(
            nombre=data['nombre'],
            fecha_inicio=datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date(),
            cantidad_inicial=int(data['cantidad_inicial']),
            cantidad_actual=int(data['cantidad_inicial']),
            genetica=data.get('genetica'),
            proveedor=data.get('proveedor'),
            peso_inicial=float(data.get('peso_inicial', 40)),
            galpon=data.get('galpon'),
            dias_ciclo=int(data.get('dias_ciclo', 42))
        )
        
        db.session.add(lote)
        db.session.commit()
        
        return jsonify({'mensaje': 'Lote creado exitosamente', 'id': lote.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al crear lote: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>', methods=['PUT'])
@token_required
def actualizar_lote(current_user, id):
    try:
        lote = Lote.query.get_or_404(id)
        data = request.get_json()
        
        print(f"🔍 DEBUG - Actualizando lote {id}")
        print(f"🔍 DEBUG - Datos recibidos: {data}")
        
        # Actualizar campos básicos
        if data.get('nombre'):
            lote.nombre = data['nombre']
        
        if data.get('fecha_inicio'):
            try:
                lote.fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date()
            except ValueError as e:
                return jsonify({'mensaje': f'Formato de fecha inválido: {str(e)}'}), 400
        
        if data.get('cantidad_inicial'):
            try:
                nueva_cantidad_inicial = int(data['cantidad_inicial'])
                cantidad_inicial_anterior = lote.cantidad_inicial
                
                # Calcular mortalidad total registrada
                total_mortalidad = sum(r.mortalidad or 0 for r in lote.registros)
                
                # Actualizar cantidad_inicial
                lote.cantidad_inicial = nueva_cantidad_inicial
                
                # Recalcular cantidad_actual basada en la nueva cantidad inicial
                lote.cantidad_actual = max(0, nueva_cantidad_inicial - total_mortalidad)
                
                print(f"🔍 DEBUG - Cantidad inicial: {cantidad_inicial_anterior} -> {nueva_cantidad_inicial}")
                print(f"🔍 DEBUG - Mortalidad total: {total_mortalidad}")
                print(f"🔍 DEBUG - Nueva cantidad actual: {lote.cantidad_actual}")
            except ValueError:
                return jsonify({'mensaje': 'Cantidad inicial debe ser un número entero'}), 400
        
        if data.get('peso_inicial'):
            try:
                lote.peso_inicial = float(data['peso_inicial'])
            except ValueError:
                return jsonify({'mensaje': 'Peso inicial debe ser un número'}), 400
        
        if data.get('genetica'):
            lote.genetica = data['genetica']
            
        if data.get('proveedor'):
            lote.proveedor = data['proveedor']
            
        if data.get('galpon'):
            lote.galpon = data['galpon']
            
        if 'dias_ciclo' in data:
            try:
                if data['dias_ciclo'] in (None, ''):
                    pass
                else:
                    lote.dias_ciclo = int(data['dias_ciclo'])
            except ValueError:
                return jsonify({'mensaje': 'Días de ciclo debe ser un número entero'}), 400
            
        if data.get('estado'):
            lote.estado = data['estado']
            # Si se está finalizando el lote, establecer fecha_fin
            if data['estado'] == 'finalizado' and not lote.fecha_fin:
                lote.fecha_fin = datetime.now().date()
        
        lote.updated_at = datetime.utcnow()
        
        print(f"✅ DEBUG - Lote actualizado: {lote.nombre}")
        db.session.commit()
        
        return jsonify({
            'mensaje': 'Lote actualizado exitosamente',
            'lote': {
                'id': lote.id,
                'nombre': lote.nombre,
                'estado': lote.estado,
                'cantidad_inicial': lote.cantidad_inicial,
                'cantidad_actual': lote.cantidad_actual
            }
        })
    except Exception as e:
        print(f"❌ DEBUG - Error actualizando lote: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'mensaje': f'Error al actualizar lote: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/cerrar', methods=['POST'])
@token_required
def cerrar_lote(current_user, id):
    try:
        lote = Lote.query.get_or_404(id)
        lote.estado = 'finalizado'
        lote.fecha_fin = datetime.now().date()
        lote.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'mensaje': 'Lote cerrado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al cerrar lote: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>', methods=['DELETE'])
@token_required
def eliminar_lote(current_user, id):
    try:
        print(f"Intentando eliminar lote {id}...")
        lote = Lote.query.get_or_404(id)
        nombre_lote = lote.nombre
        
        # Eliminar registros relacionados manualmente en el orden correcto
        try:
            # 1. Eliminar registros de sanidad (tienen FK a lote)
            num_sanidad = Sanidad.query.filter_by(lote_id=id).delete(synchronize_session=False)
            print(f"  - {num_sanidad} registros de sanidad eliminados")
            
            # 2. Eliminar registros diarios
            num_registros = RegistroDiario.query.filter_by(lote_id=id).delete(synchronize_session=False)
            print(f"  - {num_registros} registros diarios eliminados")
            
            # 3. Eliminar costos
            num_costos = Costo.query.filter_by(lote_id=id).delete(synchronize_session=False)
            print(f"  - {num_costos} costos eliminados")
            
            # 4. Eliminar ingresos
            num_ingresos = Ingreso.query.filter_by(lote_id=id).delete(synchronize_session=False)
            print(f"  - {num_ingresos} ingresos eliminados")
            
            # 5. Ahora eliminar el lote
            db.session.delete(lote)
            db.session.commit()
            print(f"✅ Lote '{nombre_lote}' eliminado exitosamente")
            
            return jsonify({'mensaje': 'Lote eliminado exitosamente'})
        except Exception as e:
            print(f"❌ Error durante la eliminación: {str(e)}")
            import traceback
            traceback.print_exc()
            db.session.rollback()
            raise
    except Exception as e:
        print(f"❌ Error al eliminar lote {id}: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar lote: {str(e)}'}), 500

# ============= RUTAS - REGISTROS DIARIOS =============

@app.route('/api/lotes/<int:id>/registros', methods=['GET'])
@token_required
def get_registros(current_user, id):
    try:
        registros = RegistroDiario.query.filter_by(lote_id=id).order_by(RegistroDiario.fecha.desc()).all()
        return jsonify([{
            'id': r.id,
            'fecha': r.fecha.isoformat(),
            'alimento_kg': r.alimento_kg,
            'agua_litros': r.agua_litros,
            'mortalidad': r.mortalidad,
            'causa_mortalidad': r.causa_mortalidad,
            'peso_promedio': r.peso_promedio,
            'temperatura_promedio': r.temperatura_promedio,
            'humedad': r.humedad,
            'observaciones': r.observaciones
        } for r in registros])
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener registros: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/registros', methods=['POST'])
@token_required
def crear_registro(current_user, id):
    try:
        data = request.get_json()
        
        # Logging para debug
        print(f"🔍 DEBUG - Creando registro para lote {id}")
        print(f"🔍 DEBUG - Datos recibidos: {data}")
        
        if not data or not data.get('fecha'):
            print("❌ DEBUG - Falta la fecha en los datos")
            return jsonify({'mensaje': 'Falta la fecha'}), 400
        
        # Verificar que el lote existe
        lote = Lote.query.get(id)
        if not lote:
            print(f"❌ DEBUG - Lote {id} no encontrado")
            return jsonify({'mensaje': 'Lote no encontrado'}), 404
        
        # Verificar si ya existe un registro para esa fecha
        try:
            fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        except (ValueError, TypeError) as e:
            print(f"❌ DEBUG - Error al convertir fecha {data['fecha']}: {e}")
            return jsonify({'mensaje': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
            
        registro_existente = RegistroDiario.query.filter_by(lote_id=id, fecha=fecha).first()
        
        if registro_existente:
            print(f"⚠️ DEBUG - Ya existe registro para lote {id} fecha {fecha}")
            return jsonify({'mensaje': 'Ya existe un registro para esta fecha'}), 400
        
        # Validar datos numéricos
        try:
            alimento_kg = float(data['alimento_kg']) if data.get('alimento_kg') not in [None, '', 0] else None
            agua_litros = float(data['agua_litros']) if data.get('agua_litros') not in [None, '', 0] else None
            mortalidad = int(data.get('mortalidad', 0))
            peso_promedio = float(data['peso_promedio']) if data.get('peso_promedio') not in [None, '', 0] else None
            temperatura_promedio = float(data['temperatura_promedio']) if data.get('temperatura_promedio') not in [None, '', 0] else None
            humedad = float(data['humedad']) if data.get('humedad') not in [None, '', 0] else None
        except (ValueError, TypeError) as e:
            print(f"❌ DEBUG - Error al convertir datos numéricos: {e}")
            return jsonify({'mensaje': f'Error en datos numéricos: {str(e)}'}), 400
        
        registro = RegistroDiario(
            lote_id=id,
            fecha=fecha,
            alimento_kg=alimento_kg,
            agua_litros=agua_litros,
            mortalidad=mortalidad,
            causa_mortalidad=data.get('causa_mortalidad'),
            peso_promedio=peso_promedio,
            temperatura_promedio=temperatura_promedio,
            humedad=humedad,
            observaciones=data.get('observaciones')
        )
        
        print(f"✅ DEBUG - Registro creado en memoria")
        db.session.add(registro)
        
        # Actualizar cantidad actual del lote si hay mortalidad
        if mortalidad > 0:
            print(f"☠️ DEBUG - Actualizando mortalidad: -{mortalidad}")
            cantidad_anterior = lote.cantidad_actual or lote.cantidad_inicial
            lote.cantidad_actual = max(0, cantidad_anterior - mortalidad)
            lote.updated_at = datetime.utcnow()
            print(f"🔢 DEBUG - Cantidad actualizada: {cantidad_anterior} -> {lote.cantidad_actual}")
        
        print(f"💾 DEBUG - Guardando en base de datos...")
        db.session.commit()
        print(f"✅ DEBUG - Registro guardado exitosamente con ID: {registro.id}")
        
        return jsonify({'mensaje': 'Registro creado exitosamente', 'id': registro.id}), 201
    except Exception as e:
        print(f"❌ DEBUG - Error: {str(e)}")
        print(f"❌ DEBUG - Tipo de error: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'mensaje': f'Error al crear registro: {str(e)}'}), 500

@app.route('/api/registros/<int:id>', methods=['PUT'])
@token_required
def actualizar_registro(current_user, id):
    try:
        registro = RegistroDiario.query.get_or_404(id)
        data = request.get_json()
        
        if data.get('alimento_kg') is not None:
            registro.alimento_kg = float(data['alimento_kg'])
        if data.get('agua_litros') is not None:
            registro.agua_litros = float(data['agua_litros'])
        if data.get('mortalidad') is not None:
            registro.mortalidad = int(data['mortalidad'])
        if data.get('causa_mortalidad') is not None:
            registro.causa_mortalidad = data['causa_mortalidad']
        if data.get('peso_promedio') is not None:
            registro.peso_promedio = float(data['peso_promedio'])
        if data.get('temperatura_promedio') is not None:
            registro.temperatura_promedio = float(data['temperatura_promedio'])
        if data.get('humedad') is not None:
            registro.humedad = float(data['humedad'])
        if data.get('observaciones') is not None:
            registro.observaciones = data['observaciones']
        
        db.session.commit()
        
        return jsonify({'mensaje': 'Registro actualizado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al actualizar registro: {str(e)}'}), 500

@app.route('/api/registros/<int:id>', methods=['DELETE'])
@token_required
def eliminar_registro(current_user, id):
    try:
        registro = RegistroDiario.query.get_or_404(id)
        db.session.delete(registro)
        db.session.commit()
        
        return jsonify({'mensaje': 'Registro eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar registro: {str(e)}'}), 500

# Endpoints específicos para registros de lotes
@app.route('/api/lotes/<int:lote_id>/registros/<int:registro_id>', methods=['GET'])
@token_required
def get_registro_lote(current_user, lote_id, registro_id):
    try:
        registro = RegistroDiario.query.filter_by(id=registro_id, lote_id=lote_id).first_or_404()
        
        return jsonify({
            'id': registro.id,
            'lote_id': registro.lote_id,
            'fecha': registro.fecha.isoformat(),
            'alimento_kg': registro.alimento_kg,
            'agua_litros': registro.agua_litros,
            'mortalidad': registro.mortalidad,
            'causa_mortalidad': registro.causa_mortalidad,
            'peso_promedio': registro.peso_promedio,
            'temperatura_promedio': registro.temperatura_promedio,
            'humedad': registro.humedad,
            'observaciones': registro.observaciones
        })
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener registro: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/registros/<int:registro_id>', methods=['PUT'])
@token_required
def actualizar_registro_lote(current_user, lote_id, registro_id):
    try:
        registro = RegistroDiario.query.filter_by(id=registro_id, lote_id=lote_id).first_or_404()
        data = request.get_json()
        
        # Actualizar campos si están presentes
        if 'fecha' in data:
            registro.fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        if 'alimento_kg' in data:
            registro.alimento_kg = float(data['alimento_kg']) if data['alimento_kg'] else None
        if 'agua_litros' in data:
            registro.agua_litros = float(data['agua_litros']) if data['agua_litros'] else None
        if 'mortalidad' in data:
            registro.mortalidad = int(data['mortalidad']) if data['mortalidad'] else 0
        if 'causa_mortalidad' in data:
            registro.causa_mortalidad = data['causa_mortalidad']
        if 'peso_promedio' in data:
            registro.peso_promedio = float(data['peso_promedio']) if data['peso_promedio'] else None
        if 'temperatura_promedio' in data:
            registro.temperatura_promedio = float(data['temperatura_promedio']) if data['temperatura_promedio'] else None
        if 'humedad' in data:
            registro.humedad = float(data['humedad']) if data['humedad'] else None
        if 'observaciones' in data:
            registro.observaciones = data['observaciones']
        
        db.session.commit()
        
        return jsonify({'mensaje': 'Registro actualizado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al actualizar registro: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/registros/<int:registro_id>', methods=['DELETE'])
@token_required
def eliminar_registro_lote(current_user, lote_id, registro_id):
    try:
        registro = RegistroDiario.query.filter_by(id=registro_id, lote_id=lote_id).first_or_404()
        db.session.delete(registro)
        db.session.commit()
        
        return jsonify({'mensaje': 'Registro eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar registro: {str(e)}'}), 500

# ============= RUTAS - ECONOMÍA =============

@app.route('/api/lotes/<int:id>/costos', methods=['GET'])
@token_required
def get_costos(current_user, id):
    try:
        costos = Costo.query.filter_by(lote_id=id).order_by(Costo.fecha.desc()).all()
        return jsonify([{
            'id': c.id,
            'categoria': c.categoria,
            'concepto': c.concepto,
            'monto': c.monto,
            'fecha': c.fecha.isoformat(),
            'observaciones': c.observaciones
        } for c in costos])
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener costos: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/costos', methods=['POST'])
@token_required
def crear_costo(current_user, id):
    try:
        data = request.get_json()
        
        if not data or not data.get('categoria') or not data.get('concepto') or not data.get('monto') or not data.get('fecha'):
            return jsonify({'mensaje': 'Faltan datos requeridos'}), 400
        
        costo = Costo(
            lote_id=id,
            categoria=data['categoria'],
            concepto=data['concepto'],
            monto=float(data['monto']),
            fecha=datetime.strptime(data['fecha'], '%Y-%m-%d').date(),
            observaciones=data.get('observaciones')
        )
        
        db.session.add(costo)
        db.session.commit()
        
        return jsonify({'mensaje': 'Costo registrado exitosamente', 'id': costo.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al registrar costo: {str(e)}'}), 500

@app.route('/api/costos/<int:id>', methods=['DELETE'])
@token_required
def eliminar_costo(current_user, id):
    try:
        costo = Costo.query.get_or_404(id)
        db.session.delete(costo)
        db.session.commit()
        
        return jsonify({'mensaje': 'Costo eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar costo: {str(e)}'}), 500

# Endpoints específicos para costos de lotes
@app.route('/api/lotes/<int:lote_id>/costos/<int:costo_id>', methods=['GET'])
@token_required
def get_costo_lote(current_user, lote_id, costo_id):
    try:
        costo = Costo.query.filter_by(id=costo_id, lote_id=lote_id).first_or_404()
        
        return jsonify({
            'id': costo.id,
            'lote_id': costo.lote_id,
            'fecha': costo.fecha.isoformat(),
            'categoria': costo.categoria,
            'concepto': costo.concepto,
            'monto': costo.monto,
            'observaciones': costo.observaciones
        })
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener costo: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/costos/<int:costo_id>', methods=['PUT'])
@token_required
def actualizar_costo_lote(current_user, lote_id, costo_id):
    try:
        costo = Costo.query.filter_by(id=costo_id, lote_id=lote_id).first_or_404()
        data = request.get_json()
        
        # Actualizar campos si están presentes
        if 'fecha' in data:
            costo.fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        if 'categoria' in data:
            costo.categoria = data['categoria']
        if 'concepto' in data:
            costo.concepto = data['concepto']
        if 'monto' in data:
            costo.monto = float(data['monto'])
        if 'observaciones' in data:
            costo.observaciones = data['observaciones']
        
        db.session.commit()
        
        return jsonify({'mensaje': 'Costo actualizado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al actualizar costo: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/costos/<int:costo_id>', methods=['DELETE'])
@token_required
def eliminar_costo_lote(current_user, lote_id, costo_id):
    try:
        costo = Costo.query.filter_by(id=costo_id, lote_id=lote_id).first_or_404()
        db.session.delete(costo)
        db.session.commit()
        
        return jsonify({'mensaje': 'Costo eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar costo: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/ingresos', methods=['GET'])
@token_required
def get_ingresos(current_user, id):
    try:
        ingresos = Ingreso.query.filter_by(lote_id=id).order_by(Ingreso.fecha.desc()).all()
        return jsonify([{
            'id': i.id,
            'cantidad_vendida': i.cantidad_vendida,
            'peso_promedio': i.peso_promedio,
            'precio_por_kg': i.precio_por_kg,
            'total': i.total,
            'fecha': i.fecha.isoformat(),
            'cliente': i.cliente,
            'observaciones': i.observaciones
        } for i in ingresos])
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener ingresos: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/ingresos', methods=['POST'])
@token_required
def crear_ingreso(current_user, id):
    try:
        data = request.get_json()
        
        if not data or not data.get('cantidad_vendida') or not data.get('peso_promedio') or not data.get('precio_por_kg') or not data.get('fecha'):
            return jsonify({'mensaje': 'Faltan datos requeridos'}), 400
        
        cantidad = int(data['cantidad_vendida'])
        peso = float(data['peso_promedio'])
        precio = float(data['precio_por_kg'])
        total = (cantidad * peso * precio) / 1000  # kg
        
        ingreso = Ingreso(
            lote_id=id,
            cantidad_vendida=cantidad,
            peso_promedio=peso,
            precio_por_kg=precio,
            total=total,
            fecha=datetime.strptime(data['fecha'], '%Y-%m-%d').date(),
            cliente=data.get('cliente'),
            observaciones=data.get('observaciones')
        )
        
        db.session.add(ingreso)
        db.session.commit()
        
        return jsonify({'mensaje': 'Ingreso registrado exitosamente', 'id': ingreso.id, 'total': total}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al registrar ingreso: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/ingresos/<int:ingreso_id>', methods=['GET'])
@token_required
def get_ingreso(current_user, lote_id, ingreso_id):
    try:
        ingreso = Ingreso.query.filter_by(id=ingreso_id, lote_id=lote_id).first_or_404()
        return jsonify({
            'id': ingreso.id,
            'cantidad_vendida': ingreso.cantidad_vendida,
            'peso_promedio': ingreso.peso_promedio,
            'precio_por_kg': ingreso.precio_por_kg,
            'total': ingreso.total,
            'fecha': ingreso.fecha.isoformat(),
            'cliente': ingreso.cliente,
            'observaciones': ingreso.observaciones
        })
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener ingreso: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/ingresos/<int:ingreso_id>', methods=['PUT'])
@token_required
def actualizar_ingreso(current_user, lote_id, ingreso_id):
    try:
        ingreso = Ingreso.query.filter_by(id=ingreso_id, lote_id=lote_id).first_or_404()
        data = request.get_json()
        
        if 'cantidad_vendida' in data:
            ingreso.cantidad_vendida = int(data['cantidad_vendida'])
        if 'peso_promedio' in data:
            ingreso.peso_promedio = float(data['peso_promedio'])
        if 'precio_por_kg' in data:
            ingreso.precio_por_kg = float(data['precio_por_kg'])
        if 'fecha' in data:
            ingreso.fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        if 'cliente' in data:
            ingreso.cliente = data['cliente']
        if 'observaciones' in data:
            ingreso.observaciones = data['observaciones']
        
        # Recalcular total
        ingreso.total = (ingreso.cantidad_vendida * ingreso.peso_promedio * ingreso.precio_por_kg) / 1000
        
        db.session.commit()
        return jsonify({'mensaje': 'Ingreso actualizado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al actualizar ingreso: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/ingresos/<int:ingreso_id>', methods=['DELETE'])
@token_required
def eliminar_ingreso(current_user, lote_id, ingreso_id):
    try:
        ingreso = Ingreso.query.filter_by(id=ingreso_id, lote_id=lote_id).first_or_404()
        db.session.delete(ingreso)
        db.session.commit()
        
        return jsonify({'mensaje': 'Ingreso eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar ingreso: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/resumen-economico', methods=['GET'])
@token_required
def get_resumen_economico(current_user, id):
    try:
        lote = Lote.query.get_or_404(id)
        
        # Agrupar costos por categoría
        costos_por_categoria = {}
        for costo in lote.costos:
            if costo.categoria not in costos_por_categoria:
                costos_por_categoria[costo.categoria] = 0
            costos_por_categoria[costo.categoria] += costo.monto
        
        total_costos = sum(c.monto for c in lote.costos)
        total_ingresos = sum(i.total for i in lote.ingresos)
        
        return jsonify({
            'total_costos': round(total_costos, 2),
            'total_ingresos': round(total_ingresos, 2),
            'ganancia': round(total_ingresos - total_costos, 2),
            'costos_por_categoria': costos_por_categoria,
            'cantidad_ventas': len(lote.ingresos)
        })
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener resumen económico: {str(e)}'}), 500

# ============= RUTAS - SANIDAD =============

@app.route('/api/lotes/<int:id>/sanidad', methods=['GET'])
@token_required
def get_sanidad(current_user, id):
    try:
        sanidad = Sanidad.query.filter_by(lote_id=id).order_by(Sanidad.fecha.desc()).all()
        return jsonify([{
            'id': s.id,
            'tipo': s.tipo,
            'producto': s.producto,
            'dosis': s.dosis,
            'fecha': s.fecha.isoformat(),
            'edad_dias': s.edad_dias,
            'via_administracion': s.via_administracion,
            'retiro_dias': s.retiro_dias,
            'enfermedad_id': s.enfermedad_id,
            'enfermedad_nombre': s.enfermedad.nombre if s.enfermedad_id else None,
            'observaciones': s.observaciones
        } for s in sanidad])
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener registros sanitarios: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/sanidad', methods=['POST'])
@token_required
def crear_sanidad(current_user, id):
    try:
        data = request.get_json()
        
        if not data or not data.get('tipo') or not data.get('producto') or not data.get('fecha'):
            return jsonify({'mensaje': 'Faltan datos requeridos'}), 400
        
        sanidad = Sanidad(
            lote_id=id,
            tipo=data['tipo'],
            producto=data['producto'],
            dosis=data.get('dosis'),
            fecha=datetime.strptime(data['fecha'], '%Y-%m-%d').date(),
            edad_dias=int(data['edad_dias']) if data.get('edad_dias') else None,
            via_administracion=data.get('via_administracion'),
            retiro_dias=int(data['retiro_dias']) if data.get('retiro_dias') else None,
            enfermedad_id=int(data['enfermedad_id']) if data.get('enfermedad_id') else None,
            observaciones=data.get('observaciones')
        )
        
        db.session.add(sanidad)
        db.session.commit()
        
        return jsonify({'mensaje': 'Registro sanitario creado exitosamente', 'id': sanidad.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al crear registro sanitario: {str(e)}'}), 500

@app.route('/api/sanidad/<int:id>', methods=['DELETE'])
@token_required
def eliminar_sanidad(current_user, id):
    try:
        sanidad = Sanidad.query.get_or_404(id)
        db.session.delete(sanidad)
        db.session.commit()
        
        return jsonify({'mensaje': 'Registro sanitario eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar registro sanitario: {str(e)}'}), 500

# Endpoints específicos para sanidad de lotes
@app.route('/api/lotes/<int:lote_id>/sanidad/<int:sanidad_id>', methods=['GET'])
@token_required
def get_sanidad_lote(current_user, lote_id, sanidad_id):
    try:
        sanidad = Sanidad.query.filter_by(id=sanidad_id, lote_id=lote_id).first_or_404()
        
        return jsonify({
            'id': sanidad.id,
            'lote_id': sanidad.lote_id,
            'fecha': sanidad.fecha.isoformat(),
            'tipo': sanidad.tipo,
            'producto': sanidad.producto,
            'dosis': sanidad.dosis,
            'edad_dias': sanidad.edad_dias,
            'via_administracion': sanidad.via_administracion,
            'retiro_dias': sanidad.retiro_dias,
            'enfermedad_id': sanidad.enfermedad_id,
            'enfermedad_nombre': sanidad.enfermedad.nombre if sanidad.enfermedad_id else None,
            'observaciones': sanidad.observaciones
        })
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener registro sanitario: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/sanidad/<int:sanidad_id>', methods=['PUT'])
@token_required
def actualizar_sanidad_lote(current_user, lote_id, sanidad_id):
    try:
        sanidad = Sanidad.query.filter_by(id=sanidad_id, lote_id=lote_id).first_or_404()
        data = request.get_json()
        
        # Actualizar campos si están presentes
        if 'fecha' in data:
            sanidad.fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        if 'tipo' in data:
            sanidad.tipo = data['tipo']
        if 'producto' in data:
            sanidad.producto = data['producto']
        if 'dosis' in data:
            sanidad.dosis = data['dosis']
        if 'edad_dias' in data:
            sanidad.edad_dias = int(data['edad_dias']) if data['edad_dias'] else None
        if 'via_administracion' in data:
            sanidad.via_administracion = data['via_administracion']
        if 'retiro_dias' in data:
            sanidad.retiro_dias = int(data['retiro_dias']) if data['retiro_dias'] else None
        if 'enfermedad_id' in data:
            sanidad.enfermedad_id = int(data['enfermedad_id']) if data['enfermedad_id'] else None
        if 'observaciones' in data:
            sanidad.observaciones = data['observaciones']
        
        db.session.commit()
        
        return jsonify({'mensaje': 'Registro sanitario actualizado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al actualizar registro sanitario: {str(e)}'}), 500

@app.route('/api/lotes/<int:lote_id>/sanidad/<int:sanidad_id>', methods=['DELETE'])
@token_required
def eliminar_sanidad_lote(current_user, lote_id, sanidad_id):
    try:
        sanidad = Sanidad.query.filter_by(id=sanidad_id, lote_id=lote_id).first_or_404()
        db.session.delete(sanidad)
        db.session.commit()
        
        return jsonify({'mensaje': 'Registro sanitario eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar registro sanitario: {str(e)}'}), 500

# ============= RUTAS - ESTADÍSTICAS Y REPORTES =============

@app.route('/api/lotes/<int:id>/estadisticas', methods=['GET'])
@token_required
def get_estadisticas(current_user, id):
    try:
        lote = Lote.query.get_or_404(id)
        estadisticas = calcular_estadisticas(lote)
        return jsonify(estadisticas)
    except Exception as e:
        return jsonify({'mensaje': f'Error al calcular estadísticas: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/curva-peso', methods=['GET'])
@token_required
def get_curva_peso(current_user, id):
    try:
        lote = Lote.query.get_or_404(id)
        registros = RegistroDiario.query.filter_by(lote_id=id).filter(
            RegistroDiario.peso_promedio.isnot(None)
        ).order_by(RegistroDiario.fecha).all()
        
        # Agregar punto inicial
        datos = [{
            'fecha': lote.fecha_inicio.isoformat(),
            'peso': lote.peso_inicial or 40,
            'dias': 0
        }]
        
        # Agregar registros
        for r in registros:
            datos.append({
                'fecha': r.fecha.isoformat(),
                'peso': r.peso_promedio,
                'dias': (r.fecha - lote.fecha_inicio).days
            })
        # ===== Curva Objetivo (simplificada Cobb 500 / Ross 308) =====
        # Valores aproximados de referencia de peso vivo promedio (g)
        curva_cobb = {
            0: 40, 1: 70, 2: 95, 3: 125, 4: 160, 5: 200, 6: 245, 7: 295,
            8: 350, 9: 410, 10: 475, 11: 545, 12: 620, 13: 700, 14: 785,
            15: 875, 16: 970, 17: 1070, 18: 1175, 19: 1285, 20: 1400,
            21: 1520, 22: 1650, 23: 1785, 24: 1925, 25: 2070, 26: 2220,
            27: 2375, 28: 2535, 29: 2700, 30: 2870, 31: 3045, 32: 3225,
            33: 3410, 34: 3600, 35: 3795, 36: 3995, 37: 4200, 38: 4410,
            39: 4625, 40: 4845, 41: 5070, 42: 5300
        }
        curva_ross = {
            0: 40, 1: 68, 2: 92, 3: 121, 4: 155, 5: 195, 6: 240, 7: 290,
            8: 344, 9: 402, 10: 465, 11: 532, 12: 603, 13: 678, 14: 757,
            15: 840, 16: 927, 17: 1018, 18: 1113, 19: 1212, 20: 1315,
            21: 1422, 22: 1533, 23: 1648, 24: 1767, 25: 1890, 26: 2017,
            27: 2148, 28: 2283, 29: 2422, 30: 2565, 31: 2712, 32: 2863,
            33: 3018, 34: 3177, 35: 3340, 36: 3507, 37: 3678, 38: 3853,
            39: 4032, 40: 4215, 41: 4402, 42: 4593
        }
        genetica = (lote.genetica or '').lower()
        es_ross = 'ross' in genetica
        curva_objetivo = curva_ross if es_ross else curva_cobb
        peso_inicial = datos[0]['peso']

        # Prepara mapa fecha->registro para consumo
        registro_por_fecha = {r.fecha.isoformat(): r for r in registros}
        consumo_acumulado = 0.0

        for i, d in enumerate(datos):
            dias = d['dias']
            d['peso_objetivo'] = curva_objetivo.get(dias)
            # ADG real (g/día)
            if i == 0:
                d['adg'] = 0
            else:
                d['adg'] = round(d['peso'] - datos[i-1]['peso'], 2)
            # Consumo acumulado (kg) y FCR
            reg = registro_por_fecha.get(d['fecha'])
            if reg and reg.alimento_kg:
                consumo_acumulado += reg.alimento_kg
            ganancia_kg = max((d['peso'] - peso_inicial) / 1000.0, 0)
            if ganancia_kg > 0:
                d['fcr'] = round(consumo_acumulado / ganancia_kg, 3)
            else:
                d['fcr'] = None
        
        return jsonify(datos)
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener curva de peso: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/curva-mortalidad', methods=['GET'])
@token_required
def get_curva_mortalidad(current_user, id):
    try:
        lote = Lote.query.get_or_404(id)
        registros = RegistroDiario.query.filter_by(lote_id=id).order_by(RegistroDiario.fecha).all()
        
        datos = []
        mortalidad_acumulada = 0
        
        for r in registros:
            mortalidad_acumulada += r.mortalidad or 0
            porcentaje = (mortalidad_acumulada / lote.cantidad_inicial * 100) if lote.cantidad_inicial > 0 else 0
            
            datos.append({
                'fecha': r.fecha.isoformat(),
                'mortalidad_diaria': r.mortalidad or 0,
                'mortalidad_acumulada': mortalidad_acumulada,
                'porcentaje': round(porcentaje, 2),
                'dias': (r.fecha - lote.fecha_inicio).days
            })
        
        return jsonify(datos)
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener curva de mortalidad: {str(e)}'}), 500

@app.route('/api/dashboard', methods=['GET'])
@token_required
def get_dashboard(current_user):
    try:
        lotes_activos = Lote.query.filter_by(estado='activo').all()
        
        dashboard_data = {
            'total_lotes_activos': len(lotes_activos),
            'total_aves': sum(l.cantidad_actual or l.cantidad_inicial for l in lotes_activos),
            'lotes': []
        }
        
        for lote in lotes_activos:
            try:
                stats = calcular_estadisticas(lote)
                dashboard_data['lotes'].append({
                    'id': lote.id,
                    'nombre': lote.nombre,
                    'galpon': lote.galpon,
                    'dias': stats['dias_transcurridos'],
                    'cantidad': stats['cantidad_actual'],
                    'peso_actual': stats['peso_actual'],
                    'mortalidad_porcentaje': stats['mortalidad_porcentaje'],
                    'fcr': stats['fcr'],
                    'adg': stats['adg'],
                    'ganancia': stats['ganancia'],
                    'rentabilidad': stats['rentabilidad']
                })
            except Exception as e:
                print(f"Error procesando lote {lote.id} en dashboard: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        return jsonify(dashboard_data)
    except Exception as e:
        print(f"Error en get_dashboard: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'mensaje': f'Error al obtener dashboard: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/alertas', methods=['GET'])
@token_required
def get_alertas(current_user, id):
    """
    Genera alertas para un lote específico basándose en el último registro diario.
    
    FÓRMULAS DETALLADAS:
    
    1. TEMPERATURA: Se compara temperatura_promedio del último registro con umbrales configurados
       - Umbral por defecto: 22°C - 32°C
       - Prioridad ALTA si desviación > 3°C del punto medio
       - Prioridad MEDIA si desviación <= 3°C
    
    2. HUMEDAD: Se compara humedad del último registro con umbrales configurados
       - Umbral por defecto: 50% - 70%
       - Prioridad ALTA si desviación > 10% del punto medio
       - Prioridad MEDIA si desviación <= 10%
    
    3. CONSUMO: Consumo diario por ave = (alimento_kg del día * 1000) / cantidad_actual_aves
       - Solo se evalúa si hay alimento registrado > 0
       - Solo genera alerta si consumo < mínimo (60g por defecto)
       - Evita falsos positivos cuando no hay registro de consumo
    
    4. MORTALIDAD DIARIA: % = (mortalidad_del_día / cantidad_aves_vivas_antes_de_ese_día) * 100
       - Se usa cantidad_actual como aproximación de aves vivas
       - Solo evalúa si mortalidad > 0 (evita división innecesaria)
       - CRÍTICO si > 2x el máximo permitido (por defecto >2%)
       - ADVERTENCIA si > máximo permitido (por defecto >1%)
       - Nota: Esta es mortalidad de UN DÍA, no acumulada
    """
    try:
        lote = Lote.query.get_or_404(id)
        stats = calcular_estadisticas(lote)
        cfg = get_configuracion_valores()
        alertas = []

        # Último registro para condiciones ambientales y consumo diario
        ultimo = RegistroDiario.query.filter_by(lote_id=id).order_by(RegistroDiario.fecha.desc()).first()
        
        # Si no hay registros, no hay alertas que generar
        if not ultimo:
            return jsonify(alertas)

        # ALERTA 1: Temperatura fuera de rango
        if ultimo.temperatura_promedio is not None and ultimo.temperatura_promedio > 0:
            temp_ideal = (cfg['temp_min'] + cfg['temp_max']) / 2
            desviacion = abs(ultimo.temperatura_promedio - temp_ideal)
            
            if ultimo.temperatura_promedio < cfg['temp_min'] or ultimo.temperatura_promedio > cfg['temp_max']:
                alertas.append({
                    'tipo': 'ADVERTENCIA',
                    'categoria': 'TEMPERATURA',
                    'mensaje': f"Temperatura fuera de rango: {ultimo.temperatura_promedio}°C (ideal {cfg['temp_min']}-{cfg['temp_max']}°C)",
                    'valor': ultimo.temperatura_promedio,
                    'prioridad': 'alta' if desviacion > 3 else 'media'
                })

        # ALERTA 2: Humedad fuera de rango
        if ultimo.humedad is not None and ultimo.humedad > 0:
            humedad_ideal = (cfg['humedad_min'] + cfg['humedad_max']) / 2
            desviacion = abs(ultimo.humedad - humedad_ideal)
            
            if ultimo.humedad < cfg['humedad_min'] or ultimo.humedad > cfg['humedad_max']:
                alertas.append({
                    'tipo': 'ADVERTENCIA',
                    'categoria': 'HUMEDAD',
                    'mensaje': f"Humedad fuera de rango: {ultimo.humedad}% (ideal {cfg['humedad_min']}-{cfg['humedad_max']}%)",
                    'valor': ultimo.humedad,
                    'prioridad': 'alta' if desviacion > 10 else 'media'
                })

        # ALERTA 3: Consumo por ave bajo (solo si hay registro de alimento)
        if ultimo.alimento_kg is not None and ultimo.alimento_kg > 0:
            cantidad_actual = stats['cantidad_actual'] or lote.cantidad_inicial
            if cantidad_actual > 0:
                # Fórmula: gramos por ave por día = (kg de alimento * 1000) / número de aves
                consumo_g_dia = (ultimo.alimento_kg * 1000) / cantidad_actual
                
                if consumo_g_dia < cfg['consumo_min_g_dia']:
                    alertas.append({
                        'tipo': 'ADVERTENCIA',
                        'categoria': 'CONSUMO',
                        'mensaje': f"Consumo bajo: {round(consumo_g_dia,1)} g/ave/día (mín {cfg['consumo_min_g_dia']} g)",
                        'valor': round(consumo_g_dia, 1),
                        'prioridad': 'media'
                    })

        # ALERTA 4: Mortalidad diaria alta (solo si hay mortalidad registrada)
        if ultimo.mortalidad is not None and ultimo.mortalidad > 0:
            # Usamos cantidad_actual como base para el cálculo
            # Nota: Idealmente debería ser la cantidad ANTES de ese día, pero cantidad_actual es aproximación
            cantidad_base = stats['cantidad_actual'] or lote.cantidad_inicial
            
            if cantidad_base > 0:
                # Fórmula: % mortalidad diaria = (aves muertas ese día / aves vivas) * 100
                pct_mortalidad_diaria = (ultimo.mortalidad / cantidad_base) * 100
                
                # Solo alertar si supera el umbral configurado
                if pct_mortalidad_diaria > cfg['mortalidad_diaria_max_pct']:
                    es_critico = pct_mortalidad_diaria > (cfg['mortalidad_diaria_max_pct'] * 2)
                    
                    alertas.append({
                        'tipo': 'CRÍTICO' if es_critico else 'ADVERTENCIA',
                        'categoria': 'MORTALIDAD',
                        'mensaje': f"Mortalidad diaria alta: {round(pct_mortalidad_diaria,2)}% (máx {cfg['mortalidad_diaria_max_pct']}%) - {ultimo.mortalidad} aves el {ultimo.fecha.strftime('%d/%m')}",
                        'valor': round(pct_mortalidad_diaria, 2),
                        'prioridad': 'alta' if es_critico else 'media'
                    })

        return jsonify(alertas)
    except Exception as e:
        print(f"Error en get_alertas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'mensaje': f'Error al obtener alertas: {str(e)}'}), 500

@app.route('/api/alertas', methods=['GET'])
@token_required
def get_alertas_generales(current_user):
    """
    Devuelve alertas de todos los lotes activos.
    Aplica las mismas fórmulas y validaciones que get_alertas() pero para múltiples lotes.
    """
    try:
        lote_id = request.args.get('lote_id', type=int)
        if lote_id:
            # Filtrar por lote específico
            return get_alertas(current_user, lote_id)

        lotes = Lote.query.filter_by(estado='activo').all()
        cfg = get_configuracion_valores()
        resultados = []
        
        for lote in lotes:
            try:
                stats = calcular_estadisticas(lote)
                ultimo = RegistroDiario.query.filter_by(lote_id=lote.id).order_by(RegistroDiario.fecha.desc()).first()
                
                # Si no hay registros para este lote, continuar con el siguiente
                if not ultimo:
                    continue
                
                # ALERTA 1: Temperatura
                if ultimo.temperatura_promedio is not None and ultimo.temperatura_promedio > 0:
                    temp_ideal = (cfg['temp_min'] + cfg['temp_max']) / 2
                    desviacion = abs(ultimo.temperatura_promedio - temp_ideal)
                    
                    if ultimo.temperatura_promedio < cfg['temp_min'] or ultimo.temperatura_promedio > cfg['temp_max']:
                        resultados.append({
                            'tipo': 'ADVERTENCIA',
                            'categoria': 'TEMPERATURA',
                            'mensaje': f"Temperatura fuera de rango: {ultimo.temperatura_promedio}°C (ideal {cfg['temp_min']}-{cfg['temp_max']}°C)",
                            'valor': ultimo.temperatura_promedio,
                            'prioridad': 'alta' if desviacion > 3 else 'media',
                            'lote_id': lote.id,
                            'lote': lote.nombre
                        })

                # ALERTA 2: Humedad
                if ultimo.humedad is not None and ultimo.humedad > 0:
                    humedad_ideal = (cfg['humedad_min'] + cfg['humedad_max']) / 2
                    desviacion = abs(ultimo.humedad - humedad_ideal)
                    
                    if ultimo.humedad < cfg['humedad_min'] or ultimo.humedad > cfg['humedad_max']:
                        resultados.append({
                            'tipo': 'ADVERTENCIA',
                            'categoria': 'HUMEDAD',
                            'mensaje': f"Humedad fuera de rango: {ultimo.humedad}% (ideal {cfg['humedad_min']}-{cfg['humedad_max']}%)",
                            'valor': ultimo.humedad,
                            'prioridad': 'alta' if desviacion > 10 else 'media',
                            'lote_id': lote.id,
                            'lote': lote.nombre
                        })

                # ALERTA 3: Consumo (solo si hay alimento registrado)
                if ultimo.alimento_kg is not None and ultimo.alimento_kg > 0:
                    cantidad_actual = stats['cantidad_actual'] or lote.cantidad_inicial
                    if cantidad_actual > 0:
                        consumo_g_dia = (ultimo.alimento_kg * 1000) / cantidad_actual
                        
                        if consumo_g_dia < cfg['consumo_min_g_dia']:
                            resultados.append({
                                'tipo': 'ADVERTENCIA',
                                'categoria': 'CONSUMO',
                                'mensaje': f"Consumo bajo: {round(consumo_g_dia,1)} g/ave/día (mín {cfg['consumo_min_g_dia']} g)",
                                'valor': round(consumo_g_dia, 1),
                                'prioridad': 'media',
                                'lote_id': lote.id,
                                'lote': lote.nombre
                            })

                # ALERTA 4: Mortalidad diaria (solo si hay mortalidad registrada)
                if ultimo.mortalidad is not None and ultimo.mortalidad > 0:
                    cantidad_base = stats['cantidad_actual'] or lote.cantidad_inicial
                    
                    if cantidad_base > 0:
                        pct_mortalidad_diaria = (ultimo.mortalidad / cantidad_base) * 100
                        
                        if pct_mortalidad_diaria > cfg['mortalidad_diaria_max_pct']:
                            es_critico = pct_mortalidad_diaria > (cfg['mortalidad_diaria_max_pct'] * 2)
                            
                            resultados.append({
                                'tipo': 'CRÍTICO' if es_critico else 'ADVERTENCIA',
                                'categoria': 'MORTALIDAD',
                                'mensaje': f"Mortalidad diaria alta: {round(pct_mortalidad_diaria,2)}% (máx {cfg['mortalidad_diaria_max_pct']}%) - {ultimo.mortalidad} aves",
                                'valor': round(pct_mortalidad_diaria, 2),
                                'prioridad': 'alta' if es_critico else 'media',
                                'lote_id': lote.id,
                                'lote': lote.nombre
                            })
                
            except Exception as e:
                print(f"Error procesando alertas para lote {lote.id}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        return jsonify(resultados)
    except Exception as e:
        print(f"Error en get_alertas_generales: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'mensaje': f'Error al obtener alertas: {str(e)}'}), 500

@app.route('/api/configuracion', methods=['GET'])
@token_required
def get_configuracion_api(current_user):
    try:
        return jsonify(get_configuracion_valores())
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener configuración: {str(e)}'}), 500

@app.route('/api/configuracion', methods=['PUT'])
@token_required
def update_configuracion_api(current_user):
    """Actualiza umbrales de configuración. Todos los campos son opcionales."""
    try:
        data = request.get_json() or {}
        cfg = Configuracion.query.first()
        if not cfg:
            cfg = Configuracion()
            db.session.add(cfg)
        campos = {
            'temp_min': float,
            'temp_max': float,
            'humedad_min': float,
            'humedad_max': float,
            'consumo_min_g_dia': float,
            'mortalidad_diaria_max_pct': float,
            'peso_tolerancia_pct': float
        }
        for campo, caster in campos.items():
            if campo in data and data[campo] is not None and data[campo] != '':
                try:
                    setattr(cfg, campo, caster(data[campo]))
                except ValueError:
                    return jsonify({'mensaje': f'Valor inválido para {campo}'}), 400
        db.session.commit()
        return jsonify({'mensaje': 'Configuración actualizada', 'configuracion': get_configuracion_valores()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al actualizar configuración: {str(e)}'}), 500

# ============= RUTAS - ENFERMEDADES =============

@app.route('/api/enfermedades', methods=['GET'])
@token_required
def listar_enfermedades(current_user):
    try:
        enfermedades = Enfermedad.query.order_by(Enfermedad.nombre.asc()).all()
        return jsonify([{
            'id': e.id,
            'nombre': e.nombre,
            'sintomas': e.sintomas,
            'prevencion': e.prevencion,
            'tratamiento': e.tratamiento,
            'medicamentos': e.medicamentos
        } for e in enfermedades])
    except Exception as e:
        return jsonify({'mensaje': f'Error al listar enfermedades: {str(e)}'}), 500

@app.route('/api/enfermedades/<int:id>', methods=['GET'])
@token_required
def obtener_enfermedad(current_user, id):
    try:
        e = Enfermedad.query.get_or_404(id)
        return jsonify({
            'id': e.id,
            'nombre': e.nombre,
            'sintomas': e.sintomas,
            'prevencion': e.prevencion,
            'tratamiento': e.tratamiento,
            'medicamentos': e.medicamentos
        })
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener enfermedad: {str(e)}'}), 500

@app.route('/api/enfermedades', methods=['POST'])
@token_required
def crear_enfermedad(current_user):
    try:
        data = request.get_json() or {}
        if not data.get('nombre'):
            return jsonify({'mensaje': 'Nombre es requerido'}), 400
        # Evitar duplicados por nombre
        if Enfermedad.query.filter_by(nombre=data['nombre']).first():
            return jsonify({'mensaje': 'Ya existe una enfermedad con ese nombre'}), 400
        e = Enfermedad(
            nombre=data['nombre'],
            sintomas=data.get('sintomas'),
            prevencion=data.get('prevencion'),
            tratamiento=data.get('tratamiento'),
            medicamentos=data.get('medicamentos')
        )
        db.session.add(e)
        db.session.commit()
        return jsonify({'mensaje': 'Enfermedad creada', 'id': e.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al crear enfermedad: {str(e)}'}), 500

@app.route('/api/enfermedades/<int:id>', methods=['PUT'])
@token_required
def actualizar_enfermedad(current_user, id):
    try:
        e = Enfermedad.query.get_or_404(id)
        data = request.get_json() or {}
        # Validar nombre duplicado si se cambia
        if 'nombre' in data and data['nombre'] and data['nombre'] != e.nombre:
            if Enfermedad.query.filter_by(nombre=data['nombre']).first():
                return jsonify({'mensaje': 'Nombre ya existe'}), 400
            e.nombre = data['nombre']
        for campo in ['sintomas', 'prevencion', 'tratamiento', 'medicamentos']:
            if campo in data:
                setattr(e, campo, data[campo])
        db.session.commit()
        return jsonify({'mensaje': 'Enfermedad actualizada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al actualizar enfermedad: {str(e)}'}), 500

@app.route('/api/enfermedades/<int:id>', methods=['DELETE'])
@token_required
def eliminar_enfermedad(current_user, id):
    try:
        e = Enfermedad.query.get_or_404(id)
        db.session.delete(e)
        db.session.commit()
        return jsonify({'mensaje': 'Enfermedad eliminada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al eliminar enfermedad: {str(e)}'}), 500

# ============= RUTA - QR POR LOTE =============
@app.route('/api/lotes/<int:id>/qr', methods=['GET'])
@token_required
def generar_qr_lote(current_user, id):
    """Genera un PNG con un QR que contiene una URL pública del lote.
    La URL base puede configurarse con PUBLIC_LOTE_BASE_URL; por defecto, ejemplo genérico."""
    if not QRCODE_AVAILABLE:
        return jsonify({
            'mensaje': 'Dependencia qrcode no instalada',
            'detalle': 'Ejecutar: pip install qrcode[pil]'
        }), 500
    
    try:
        lote = Lote.query.get_or_404(id)
        # Construir URL pública: si existe env var se usa, si no se usa la propia página pública local
        base_public_url = os.environ.get('PUBLIC_LOTE_BASE_URL')
        if base_public_url:
            payload_url = f"{base_public_url.rstrip('/')}/{lote.id}"
        else:
            # Usar host actual para la página pública HTML
            payload_url = request.host_url.rstrip('/') + f"/public/lote/{lote.id}"
        
        # Generar QR
        img = qrcode.make(payload_url)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return send_file(buf, mimetype='image/png', download_name=f'lote_{lote.id}_qr.png')
    except Exception as e:
        print(f"Error generando QR para lote {id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'mensaje': f'Error al generar QR: {str(e)}'}), 500

# ============= RUTA PÚBLICA - RESUMEN DE LOTE =============
@app.route('/api/public/lotes/<int:id>', methods=['GET'])
def obtener_lote_publico(id):
    """Endpoint público (sin token) que expone un resumen del lote.
    No incluye datos sensibles ni costos/ingresos detallados."""
    try:
        lote = Lote.query.get_or_404(id)
        stats = calcular_estadisticas(lote)
        resumen = {
            'id': lote.id,
            'nombre': lote.nombre,
            'fecha_inicio': lote.fecha_inicio.isoformat(),
            'fecha_sacrificio': (lote.fecha_inicio + timedelta(days=lote.dias_ciclo or 42)).isoformat(),
            'dias_ciclo': lote.dias_ciclo or 42,
            'dias_transcurridos': stats.get('dias_transcurridos'),
            'dias_restantes': (lote.dias_ciclo or 42) - stats.get('dias_transcurridos', 0),
            'genetica': lote.genetica,
            'cantidad_inicial': lote.cantidad_inicial,
            'cantidad_actual': stats.get('cantidad_actual'),
            'peso_promedio_actual_g': stats.get('peso_actual'),
            'fcr': stats.get('fcr'),
            'adg_g_dia': stats.get('adg'),
            'mortalidad_porcentaje': stats.get('mortalidad_porcentaje')
        }
        return jsonify(resumen)
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener lote público: {str(e)}'}), 500

@app.route('/public/lote/<int:id>', methods=['GET'])
def pagina_publica_lote(id):
        """Página pública simple en HTML que consume el JSON del lote."""
        try:
                # Verificar existencia del lote para responder 404 si no existe
                Lote.query.get_or_404(id)
                # HTML minimalista con fetch al endpoint público JSON
                html = f"""<!DOCTYPE html>
<html lang='es'>
<head>
    <meta charset='utf-8'>
    <meta name='viewport' content='width=device-width, initial-scale=1'>
    <title>Reporte público del lote #{id}</title>
    <style>
        :root {{
            --bg: #0f172a; --card:#0b1220; --text:#e2e8f0; --muted:#94a3b8; --border:#334155; --accent:#2563eb;
        }}
        body {{ margin:0; font-family: system-ui, -apple-system, Segoe UI, Roboto, sans-serif; background: var(--bg); color: var(--text); }}
        .container {{ max-width: 720px; margin: 0 auto; padding: 1.25rem; }}
        .card {{ background: var(--card); border:1px solid var(--border); border-radius:12px; padding:1.25rem; box-shadow: 0 1px 3px rgba(0,0,0,.5); }}
        h1 {{ font-size:1.5rem; margin: 0 0 0.75rem; color: var(--text); }}
        .muted {{ color: var(--muted); }}
        .grid {{ display:grid; grid-template-columns: repeat(auto-fit, minmax(200px,1fr)); gap:.75rem; margin-top: .75rem; }}
        .stat {{ background:#0f172a; border:1px solid var(--border); border-radius:10px; padding:.75rem; }}
        .label {{ font-size:.75rem; color:var(--muted); text-transform:uppercase; letter-spacing:.04em; }}
        .value {{ font-size:1.1rem; font-weight:600; }}
        .header {{ display:flex; align-items:center; justify-content:space-between; margin-bottom: .75rem; }}
        a.btn {{ display:inline-block; background:var(--accent); color:white; padding:.5rem .75rem; border-radius:8px; text-decoration:none; font-weight:600; }}
    </style>
    <script>
        async function load() {{
            try {{
                const res = await fetch('/api/public/lotes/{id}');
                if (!res.ok) throw new Error('No encontrado');
                const d = await res.json();
                document.getElementById('title').textContent = `Lote #${{d.id}} · ${{d.nombre || ''}}`;
                document.getElementById('inicio').textContent = new Date(d.fecha_inicio).toLocaleDateString('es-CO');
                document.getElementById('sacrificio').textContent = new Date(d.fecha_sacrificio).toLocaleDateString('es-CO');
                document.getElementById('dias').textContent = `${{d.dias_transcurridos}} / ${{d.dias_ciclo}} días`;
                document.getElementById('restantes').textContent = d.dias_restantes;
                document.getElementById('genetica').textContent = d.genetica || '-';
                document.getElementById('aves').textContent = `${{d.cantidad_actual}} / ${{d.cantidad_inicial}}`;
                document.getElementById('peso').textContent = (d.peso_promedio_actual_g ?? '-') + ' g';
                document.getElementById('fcr').textContent = d.fcr ?? '-';
                document.getElementById('adg').textContent = (d.adg_g_dia ?? '-') + ' g/día';
                document.getElementById('mort').textContent = (d.mortalidad_porcentaje ?? 0) + '%';
            }} catch(e) {{
                document.getElementById('root').innerHTML = '<div class="card"><p>⚠️ Lote no disponible o enlace inválido.</p></div>';
            }}
        }}
        addEventListener('DOMContentLoaded', load);
    </script>
</head>
<body>
    <div class='container'>
        <div class='card'>
            <div class='header'>
                <h1 id='title'>Lote #{id}</h1>
                <a class='btn' href='/api/public/lotes/{id}' target='_blank' rel='noopener'>Ver JSON</a>
            </div>
            <div class='grid'>
                <div class='stat'><div class='label'>Inicio</div><div class='value' id='inicio'>-</div></div>
                <div class='stat'><div class='label'>Sacrificio</div><div class='value' id='sacrificio'>-</div></div>
                <div class='stat'><div class='label'>Ciclo</div><div class='value' id='dias'>-</div></div>
                <div class='stat'><div class='label'>Restantes</div><div class='value' id='restantes'>-</div></div>
                <div class='stat'><div class='label'>Genética</div><div class='value' id='genetica'>-</div></div>
                <div class='stat'><div class='label'>Aves</div><div class='value' id='aves'>-</div></div>
                <div class='stat'><div class='label'>Peso Prom.</div><div class='value' id='peso'>-</div></div>
                <div class='stat'><div class='label'>FCR</div><div class='value' id='fcr'>-</div></div>
                <div class='stat'><div class='label'>ADG</div><div class='value' id='adg'>-</div></div>
                <div class='stat'><div class='label'>Mortalidad</div><div class='value' id='mort'>-</div></div>
            </div>
            <p class='muted' style='margin-top:.75rem;'>Enlace generado automáticamente. Compartible sin iniciar sesión.</p>
        </div>
    </div>
</body>
</html>"""
                return app.response_class(html, mimetype='text/html')
        except Exception as e:
                return jsonify({'mensaje': f'Error al cargar página pública del lote: {str(e)}'}), 500

# ============= RUTAS - COMPARACIÓN Y ANÁLISIS =============

@app.route('/api/comparar-lotes', methods=['POST'])
@token_required
def comparar_lotes(current_user):
    try:
        data = request.get_json()
        lote_ids = data.get('lotes', [])
        
        if not lote_ids or len(lote_ids) < 2:
            return jsonify({'mensaje': 'Se requieren al menos 2 lotes para comparar'}), 400
        
        comparacion = []
        
        for lote_id in lote_ids:
            lote = Lote.query.get(lote_id)
            if lote:
                stats = calcular_estadisticas(lote)
                comparacion.append({
                    'id': lote.id,
                    'nombre': lote.nombre,
                    'fcr': stats['fcr'],
                    'adg': stats['adg'],
                    'mortalidad_porcentaje': stats['mortalidad_porcentaje'],
                    'rentabilidad': stats['rentabilidad'],
                    'costo_por_kg': stats['costo_por_kg'],
                    'dias_transcurridos': stats['dias_transcurridos']
                })
        
        return jsonify(comparacion)
    except Exception as e:
        return jsonify({'mensaje': f'Error al comparar lotes: {str(e)}'}), 500

@app.route('/api/estadisticas-generales', methods=['GET'])
@token_required
def get_estadisticas_generales(current_user):
    try:
        lotes = Lote.query.all()
        
        if not lotes:
            return jsonify({
                'total_lotes': 0,
                'lotes_activos': 0,
                'lotes_finalizados': 0,
                'total_aves_procesadas': 0
            })
        
        lotes_activos = [l for l in lotes if l.estado == 'activo']
        lotes_finalizados = [l for l in lotes if l.estado == 'finalizado']
        
        # Promedios de lotes finalizados
        if lotes_finalizados:
            stats_finalizados = [calcular_estadisticas(l) for l in lotes_finalizados]
            fcr_promedio = sum(s['fcr'] for s in stats_finalizados) / len(stats_finalizados)
            mortalidad_promedio = sum(s['mortalidad_porcentaje'] for s in stats_finalizados) / len(stats_finalizados)
            rentabilidad_promedio = sum(s['rentabilidad'] for s in stats_finalizados) / len(stats_finalizados)
        else:
            fcr_promedio = 0
            mortalidad_promedio = 0
            rentabilidad_promedio = 0
        
        return jsonify({
            'total_lotes': len(lotes),
            'lotes_activos': len(lotes_activos),
            'lotes_finalizados': len(lotes_finalizados),
            'total_aves_procesadas': sum(l.cantidad_inicial for l in lotes),
            'fcr_promedio': round(fcr_promedio, 2),
            'mortalidad_promedio': round(mortalidad_promedio, 2),
            'rentabilidad_promedio': round(rentabilidad_promedio, 2)
        })
    except Exception as e:
        return jsonify({'mensaje': f'Error al obtener estadísticas generales: {str(e)}'}), 500

# ============= RUTAS - EXPORTACIÓN =============

@app.route('/api/lotes/<int:id>/export', methods=['GET'])
@token_required
def exportar_lote_semana(current_user, id):
    """Exporta un resumen semanal del lote en PDF o Excel.
    Parámetros:
      - formato: pdf|xlsx (por defecto: pdf)
      - semana: YYYY-WW o YYYY-Www (ISO week). Si no se pasa, usa la semana actual del ciclo desde fecha_inicio.
    """
    try:
        lote = Lote.query.get_or_404(id)
        formato = (request.args.get('formato') or 'pdf').lower()
        semana_param = request.args.get('semana')

        # Calcular rango de semana
        if semana_param:
            try:
                year_str, week_str = semana_param.split('-')
                if week_str and week_str[0] in ('W', 'w'):
                    week_str = week_str[1:]
                start_date = date.fromisocalendar(int(year_str), int(week_str), 1)
                end_date = start_date + timedelta(days=6)
            except Exception:
                return jsonify({'mensaje': 'Parámetro semana inválido. Use YYYY-WW'}), 400
        else:
            hoy = datetime.now().date()
            dias_transc = max((hoy - lote.fecha_inicio).days, 0)
            semana_idx = dias_transc // 7
            start_date = lote.fecha_inicio + timedelta(days=semana_idx * 7)
            end_date = start_date + timedelta(days=6)

        # Consultar datos de la semana
        registros = RegistroDiario.query.filter(
            RegistroDiario.lote_id == id,
            RegistroDiario.fecha >= start_date,
            RegistroDiario.fecha <= end_date
        ).order_by(RegistroDiario.fecha.asc()).all()

        costos = Costo.query.filter(
            Costo.lote_id == id,
            Costo.fecha >= start_date,
            Costo.fecha <= end_date
        ).order_by(Costo.fecha.asc()).all()

        sanidad_semana = Sanidad.query.filter(
            Sanidad.lote_id == id,
            Sanidad.fecha >= start_date,
            Sanidad.fecha <= end_date
        ).order_by(Sanidad.fecha.asc()).all()

        # Resumen
        alimento = sum((r.alimento_kg or 0) for r in registros)
        agua = sum((r.agua_litros or 0) for r in registros)
        mort = sum((r.mortalidad or 0) for r in registros)
        temp_prom = round(sum((r.temperatura_promedio or 0) for r in registros) / len(registros), 2) if registros else 0
        hum_prom = round(sum((r.humedad or 0) for r in registros) / len(registros), 2) if registros else 0
        peso_ultimo = None
        for r in reversed(registros):
            if r.peso_promedio is not None:
                peso_ultimo = r.peso_promedio
                break
        costos_total = round(sum(c.monto for c in costos), 2)

        if formato in ('pdf', 'application/pdf'):
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors
            except Exception as e:
                return jsonify({'mensaje': 'reportlab no instalado', 'detalle': str(e)}), 500

            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4)
            styles = getSampleStyleSheet()
            flow = []
            flow.append(Paragraph(f"Reporte semanal · Lote #{lote.id} · {lote.nombre}", styles['Title']))
            flow.append(Paragraph(f"Semana: {start_date.isoformat()} a {end_date.isoformat()}", styles['Normal']))
            flow.append(Spacer(1, 12))

            resumen_data = [
                ['Aves actuales', lote.cantidad_actual or lote.cantidad_inicial, 'Alimento (kg)', round(alimento, 2)],
                ['Mortalidad', mort, 'Agua (L)', round(agua, 2)],
                ['Temp prom (°C)', temp_prom, 'Humedad prom (%)', hum_prom],
                ['Peso prom último (g)', peso_ultimo or '-', 'Costos ($)', costos_total]
            ]
            tabla_resumen = Table(resumen_data, colWidths=[120, 120, 120, 120])
            tabla_resumen.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            flow.append(tabla_resumen)
            flow.append(Spacer(1, 12))

            reg_data = [['Fecha', 'Alimento (kg)', 'Agua (L)', 'Mort', 'Peso (g)', 'Temp', 'Hum (%)']]
            for r in registros[:20]:
                reg_data.append([
                    r.fecha.isoformat(),
                    r.alimento_kg or '-',
                    r.agua_litros or '-',
                    r.mortalidad or 0,
                    r.peso_promedio or '-',
                    r.temperatura_promedio or '-',
                    r.humedad or '-'
                ])
            if len(reg_data) > 1:
                t = Table(reg_data)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ]))
                flow.append(Paragraph('Registros de la semana', styles['Heading3']))
                flow.append(t)
                flow.append(Spacer(1, 12))

            cost_data = [['Fecha', 'Categoría', 'Concepto', 'Monto']]
            for c in costos[:20]:
                cost_data.append([c.fecha.isoformat(), c.categoria, c.concepto, f"{c.monto:,.0f}"])
            if len(cost_data) > 1:
                t2 = Table(cost_data, colWidths=[80, 90, 220, 70])
                t2.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ]))
                flow.append(Paragraph('Costos de la semana', styles['Heading3']))
                flow.append(t2)

            if sanidad_semana:
                sanidad_data = [['Fecha','Tipo','Producto','Dosis','Edad','Vía','Enfermedad']]
                for s in sanidad_semana[:15]:
                    sanidad_data.append([
                        s.fecha.isoformat(), s.tipo, s.producto,
                        s.dosis or '-', s.edad_dias or '-', s.via_administracion or '-', s.enfermedad.nombre if s.enfermedad_id else '-'
                    ])
                t3 = Table(sanidad_data, colWidths=[70,55,160,60,40,70,90])
                t3.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                    ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ]))
                flow.append(Spacer(1,12))
                flow.append(Paragraph('Sanidad de la semana', styles['Heading3']))
                flow.append(t3)

            # Gráfico peso vs objetivo (semana) simple si hay datos
            try:
                from reportlab.graphics.shapes import Drawing, Line, String
                from reportlab.graphics import renderPDF
                registros_peso = [r for r in registros if r.peso_promedio is not None]
                if registros_peso:
                    curva_cobb = {0:40,1:70,2:95,3:125,4:160,5:200,6:245,7:295,8:350,9:410,10:475,11:545,12:620,13:700,14:785,15:875,16:970,17:1070,18:1175,19:1285,20:1400,21:1520,22:1650,23:1785,24:1925,25:2070,26:2220,27:2375,28:2535,29:2700,30:2870,31:3045,32:3225,33:3410,34:3600,35:3795,36:3995,37:4200,38:4410,39:4625,40:4845,41:5070,42:5300}
                    curva_ross = {0:40,1:68,2:92,3:121,4:155,5:195,6:240,7:290,8:344,9:402,10:465,11:532,12:603,13:678,14:757,15:840,16:927,17:1018,18:1113,19:1212,20:1315,21:1422,22:1533,23:1648,24:1767,25:1890,26:2017,27:2148,28:2283,29:2422,30:2565,31:2712,32:2863,33:3018,34:3177,35:3340,36:3507,37:3678,38:3853,39:4032,40:4215,41:4402,42:4593}
                    es_ross = 'ross' in (lote.genetica or '').lower()
                    curva_obj = curva_ross if es_ross else curva_cobb
                    puntos = []
                    for r in registros_peso:
                        dias = (r.fecha - lote.fecha_inicio).days
                        puntos.append((dias, r.peso_promedio, curva_obj.get(dias)))
                    if puntos:
                        dias_vals = [p[0] for p in puntos]
                        peso_vals = [p[1] for p in puntos]
                        obj_vals = [p[2] for p in puntos if p[2] is not None]
                        min_d, max_d = min(dias_vals), max(dias_vals)
                        min_p = min(peso_vals + obj_vals)
                        max_p = max(peso_vals + obj_vals)
                        W, H = 420, 180
                        d = Drawing(W, H)
                        d.add(Line(40,30,40,H-20))
                        d.add(Line(40,30,W-20,30))
                        def sx(x):
                            return 40 + ((x - min_d) / (max_d - min_d or 1)) * (W-60)
                        def sy(y):
                            return 30 + ((y - min_p) / (max_p - min_p or 1)) * (H-60)
                        for i in range(1, len(puntos)):
                            x1, p1, o1 = puntos[i-1]
                            x2, p2, o2 = puntos[i]
                            d.add(Line(sx(x1), sy(p1), sx(x2), sy(p2), strokeColor=colors.HexColor('#2563eb'), strokeWidth=2))
                            if o1 is not None and o2 is not None:
                                d.add(Line(sx(x1), sy(o1), sx(x2), sy(o2), strokeColor=colors.HexColor('#16a34a'), strokeWidth=1.2))
                        d.add(String(45, H-25, 'Peso vs Objetivo (semana)', fontSize=10))
                        renderPDF.draw(d, doc.canv, 70, 100)
            except Exception:
                pass

            doc.build(flow)
            buf.seek(0)
            filename = f"lote_{lote.id}_semana_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
            return send_file(buf, mimetype='application/pdf', download_name=filename)

        elif formato in ('xlsx', 'excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
            except Exception as e:
                return jsonify({'mensaje': 'openpyxl no instalado', 'detalle': str(e)}), 500

            buf = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Resumen'
            ws.append(['Lote', f"#{lote.id} - {lote.nombre}"])
            ws.append(['Periodo', f"{start_date} a {end_date}"])
            ws.append(['Aves actuales', lote.cantidad_actual or lote.cantidad_inicial])
            ws.append(['Alimento (kg)', round(alimento, 2)])
            ws.append(['Agua (L)', round(agua, 2)])
            ws.append(['Mortalidad', mort])
            ws.append(['Temp prom (°C)', temp_prom])
            ws.append(['Humedad prom (%)', hum_prom])
            ws.append(['Peso prom último (g)', peso_ultimo or '-'])
            ws.append(['Costos ($)', costos_total])

            ws2 = wb.create_sheet('Registros')
            ws2.append(['Fecha', 'Alimento (kg)', 'Agua (L)', 'Mortalidad', 'Peso (g)', 'Temp (°C)', 'Humedad (%)', 'Observaciones'])
            for r in registros:
                ws2.append([
                    r.fecha.isoformat(), r.alimento_kg, r.agua_litros, r.mortalidad, r.peso_promedio,
                    r.temperatura_promedio, r.humedad, (r.observaciones or '')[:200]
                ])

            ws3 = wb.create_sheet('Costos')
            ws3.append(['Fecha', 'Categoría', 'Concepto', 'Monto'])
            for c in costos:
                ws3.append([c.fecha.isoformat(), c.categoria, c.concepto, c.monto])

            ws4 = wb.create_sheet('Sanidad')
            ws4.append(['Fecha','Tipo','Producto','Dosis','Edad (días)','Vía','Enfermedad','Retiro (días)'])
            for s in sanidad_semana:
                ws4.append([
                    s.fecha.isoformat(), s.tipo, s.producto, s.dosis, s.edad_dias,
                    s.via_administracion, s.enfermedad.nombre if s.enfermedad_id else '', s.retiro_dias
                ])

            for sheet in (ws, ws2, ws3, ws4):
                for col in sheet.columns:
                    max_len = 0
                    letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            max_len = max(max_len, len(str(cell.value)))
                        except Exception:
                            pass
                    sheet.column_dimensions[letter].width = min(max_len + 2, 50)

            from openpyxl.styles import PatternFill, Font
            header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            for sheet in (ws, ws2, ws3, ws4):
                for cell in sheet[1]:  # primera fila
                    cell.fill = header_fill
                    cell.font = header_font

            wb.save(buf)
            buf.seek(0)
            filename = f"lote_{lote.id}_semana_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name=filename)

        else:
            return jsonify({'mensaje': 'Formato no soportado. Use pdf o xlsx'}), 400

    except Exception as e:
        return jsonify({'mensaje': f'Error al exportar: {str(e)}'}), 500

@app.route('/api/lotes/<int:id>/exportar-csv', methods=['GET'])
@token_required
def exportar_csv(current_user, id):
    try:
        lote = Lote.query.get_or_404(id)
        registros = RegistroDiario.query.filter_by(lote_id=id).order_by(RegistroDiario.fecha).all()
        
        # Crear CSV en memoria
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow(['Fecha', 'Alimento (kg)', 'Agua (L)', 'Mortalidad', 'Peso (g)', 'Temperatura (°C)', 'Observaciones'])
        
        # Datos
        for r in registros:
            writer.writerow([
                r.fecha.isoformat(),
                r.alimento_kg or '',
                r.agua_litros or '',
                r.mortalidad or 0,
                r.peso_promedio or '',
                r.temperatura_promedio or '',
                r.observaciones or ''
            ])
        
        output.seek(0)
        
        return output.getvalue(), 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename=lote_{lote.nombre.replace(" ", "_")}_registros.csv'
        }
    except Exception as e:
        return jsonify({'mensaje': f'Error al exportar CSV: {str(e)}'}), 500

# ============= INICIALIZACIÓN =============

@app.route('/api/init', methods=['POST'])
def init_database():
    """Inicializa la base de datos con un usuario por defecto"""
    try:
        db.create_all()
        
        # Verificar si ya existe un usuario
        if Usuario.query.first():
            return jsonify({'mensaje': 'Base de datos ya inicializada'}), 400
        
        # Crear usuario por defecto
        usuario = Usuario(
            username='admin',
            password_hash=generate_password_hash('admin123'),
            nombre_completo='Administrador',
            email='admin@granja.com',
            rol='admin'
        )
        
        db.session.add(usuario)
        db.session.commit()
        
        return jsonify({
            'mensaje': 'Base de datos inicializada exitosamente',
            'usuario': 'admin',
            'password': 'admin123',
            'advertencia': 'Cambiar contraseña en producción'
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'Error al inicializar: {str(e)}'}), 500

@app.route('/', methods=['GET'])
def serve_frontend():
    """Sirve la SPA del frontend (index.html)."""
    try:
        return app.send_static_file('index.html')
    except Exception:
        return jsonify({'mensaje': 'Frontend no encontrado', 'ruta': FRONTEND_DIR}), 404

@app.route('/api', methods=['GET'])
def api_info():
    return jsonify({
        'mensaje': 'API Sistema Control de Pollos de Engorde',
        'version': '1.0',
        'estado': 'activo',
        'endpoints': {
            'auth': '/api/auth/login, /api/auth/register',
            'lotes': '/api/lotes',
            'registros': '/api/lotes/:id/registros',
            'costos': '/api/lotes/:id/costos',
            'sanidad': '/api/lotes/:id/sanidad',
            'estadisticas': '/api/lotes/:id/estadisticas',
            'dashboard': '/api/dashboard',
            'alertas': '/api/alertas?lote_id=:id',
            'configuracion': '/api/configuracion',
            'enfermedades': '/api/enfermedades'
        }
    })



# ============= ERROR HANDLERS =============

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Recurso no encontrado', 'codigo': 404}), 404

@app.errorhandler(400)
def bad_request(error):
    return jsonify({'error': 'Petición incorrecta', 'codigo': 400}), 400

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return jsonify({'error': 'Error interno del servidor', 'codigo': 500}), 500

@app.errorhandler(Exception)
def handle_exception(error):
    db.session.rollback()
    return jsonify({
        'error': 'Error inesperado',
        'mensaje': str(error),
        'codigo': 500
    }), 500

# ============= ENDPOINT DE ESTADO =============

@app.route('/api/health', methods=['GET'])
def health_check():
    """Endpoint para verificar el estado del servidor"""
    try:
        # Verificar conexión a BD
        db.session.execute('SELECT 1')
        
        # Contar registros básicos
        usuarios_count = Usuario.query.count()
        lotes_count = Lote.query.count()
        registros_count = RegistroDiario.query.count()
        
        return jsonify({
            'estado': 'OK',
            'mensaje': 'Servidor funcionando correctamente',
            'timestamp': datetime.now().isoformat(),
            'base_datos': {
                'conectado': True,
                'usuarios': usuarios_count,
                'lotes': lotes_count,
                'registros': registros_count
            },
            'version': '1.0.0'
        }), 200
        
    except Exception as e:
        return jsonify({
            'estado': 'ERROR',
            'mensaje': f'Error en servidor: {str(e)}',
            'timestamp': datetime.now().isoformat(),
            'base_datos': {'conectado': False}
        }), 500

@app.route('/api/test', methods=['POST'])
def test_endpoint():
    """Endpoint de prueba para validar datos"""
    try:
        data = request.get_json()
        return jsonify({
            'mensaje': 'Datos recibidos correctamente',
            'datos_recibidos': data,
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'error': 'Error procesando datos de prueba',
            'mensaje': str(e)
        }), 400

# ============= EJECUCIÓN =============

if __name__ == '__main__':
    # En entorno local, arrancar con servidor de desarrollo
    port = int(os.environ.get('PORT', 5000))
    host = os.environ.get('HOST', '0.0.0.0')
    print(f"\n🚀 Servidor DEV en http://{host}:{port}")
    app.run(debug=True, host=host, port=port)
