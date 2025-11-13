"""
Servicio para exportar datos a PDF y Excel
"""
import os
from datetime import datetime
from io import BytesIO
import csv

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExportadorPDF:
    """Exporta datos de lotes a PDF"""
    
    def __init__(self):
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab no está instalado. pip install reportlab")
        self.styles = getSampleStyleSheet()
        
    def generar_reporte_lote(self, lote_data: dict, registros: list, 
                            costos: list, estadisticas: dict, filename: str):
        """Genera un reporte completo del lote en PDF"""
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=30,
        )
        title = Paragraph(f"Reporte de Lote: {lote_data['nombre']}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Información general
        info_data = [
            ['Información General', ''],
            ['Fecha de inicio:', lote_data['fecha_inicio']],
            ['Galpón:', lote_data.get('galpon', 'N/A')],
            ['Genética:', lote_data.get('genetica', 'N/A')],
            ['Cantidad inicial:', str(lote_data['cantidad_inicial'])],
            ['Cantidad actual:', str(estadisticas.get('cantidad_actual', 0))],
            ['Días transcurridos:', str(estadisticas.get('dias_transcurridos', 0))],
        ]
        
        info_table = Table(info_data, colWidths=[3*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Estadísticas de rendimiento
        stats_data = [
            ['Indicadores de Rendimiento', ''],
            ['FCR (Conversión Alimenticia):', f"{estadisticas.get('fcr', 0):.2f}"],
            ['ADG (Ganancia Diaria):', f"{estadisticas.get('adg', 0):.2f} g/día"],
            ['Mortalidad:', f"{estadisticas.get('mortalidad_porcentaje', 0):.2f}%"],
            ['Peso actual promedio:', f"{estadisticas.get('peso_actual', 0):.2f} g"],
            ['Alimento total consumido:', f"{estadisticas.get('total_alimento', 0):.2f} kg"],
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Análisis económico
        econ_data = [
            ['Análisis Económico', ''],
            ['Total costos:', f"${estadisticas.get('total_costos', 0):,.0f}"],
            ['Total ingresos:', f"${estadisticas.get('total_ingresos', 0):,.0f}"],
            ['Ganancia:', f"${estadisticas.get('ganancia', 0):,.0f}"],
            ['Rentabilidad:', f"{estadisticas.get('rentabilidad', 0):.2f}%"],
            ['Costo por kg:', f"${estadisticas.get('costo_por_kg', 0):,.0f}"],
            ['Costo por pollo:', f"${estadisticas.get('costo_por_pollo', 0):,.0f}"],
        ]
        
        econ_table = Table(econ_data, colWidths=[3*inch, 3*inch])
        econ_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(econ_table)
        
        # Generar PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Guardar archivo
        with open(filename, 'wb') as f:
            f.write(buffer.getvalue())
        
        return filename

class ExportadorExcel:
    """Exporta datos de lotes a Excel"""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. pip install openpyxl")
    
    def generar_reporte_lote(self, lote_data: dict, registros: list, 
                            costos: list, sanidad: list, estadisticas: dict, 
                            filename: str):
        """Genera un reporte completo del lote en Excel"""
        
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        self._crear_hoja_resumen(ws_resumen, lote_data, estadisticas)
        
        # Hoja 2: Registros diarios
        ws_registros = wb.create_sheet("Registros Diarios")
        self._crear_hoja_registros(ws_registros, registros)
        
        # Hoja 3: Costos
        ws_costos = wb.create_sheet("Costos")
        self._crear_hoja_costos(ws_costos, costos)
        
        # Hoja 4: Sanidad
        ws_sanidad = wb.create_sheet("Sanidad")
        self._crear_hoja_sanidad(ws_sanidad, sanidad)
        
        # Guardar archivo
        wb.save(filename)
        return filename
    
    def _crear_hoja_resumen(self, ws, lote_data, estadisticas):
        """Crea la hoja de resumen"""
        
        # Estilos
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)
        
        # Título
        ws['A1'] = f"Reporte de Lote: {lote_data['nombre']}"
        ws['A1'].font = Font(size=18, bold=True)
        ws.merge_cells('A1:B1')
        
        # Información general
        ws['A3'] = "INFORMACIÓN GENERAL"
        ws['A3'].fill = header_fill
        ws['A3'].font = header_font
        ws.merge_cells('A3:B3')
        
        info = [
            ('Fecha de inicio:', lote_data['fecha_inicio']),
            ('Galpón:', lote_data.get('galpon', 'N/A')),
            ('Genética:', lote_data.get('genetica', 'N/A')),
            ('Cantidad inicial:', lote_data['cantidad_inicial']),
            ('Cantidad actual:', estadisticas.get('cantidad_actual', 0)),
            ('Días transcurridos:', estadisticas.get('dias_transcurridos', 0)),
        ]
        
        row = 4
        for label, value in info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Indicadores de rendimiento
        ws[f'A{row+1}'] = "INDICADORES DE RENDIMIENTO"
        ws[f'A{row+1}'].fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        ws[f'A{row+1}'].font = header_font
        ws.merge_cells(f'A{row+1}:B{row+1}')
        
        stats = [
            ('FCR:', estadisticas.get('fcr', 0)),
            ('ADG (g/día):', estadisticas.get('adg', 0)),
            ('Mortalidad (%):', estadisticas.get('mortalidad_porcentaje', 0)),
            ('Peso actual (g):', estadisticas.get('peso_actual', 0)),
            ('Alimento total (kg):', estadisticas.get('total_alimento', 0)),
        ]
        
        row += 2
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Análisis económico
        ws[f'A{row+1}'] = "ANÁLISIS ECONÓMICO"
        ws[f'A{row+1}'].fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        ws[f'A{row+1}'].font = header_font
        ws.merge_cells(f'A{row+1}:B{row+1}')
        
        econ = [
            ('Total costos:', f"${estadisticas.get('total_costos', 0):,.0f}"),
            ('Total ingresos:', f"${estadisticas.get('total_ingresos', 0):,.0f}"),
            ('Ganancia:', f"${estadisticas.get('ganancia', 0):,.0f}"),
            ('Rentabilidad (%):', estadisticas.get('rentabilidad', 0)),
            ('Costo por kg:', f"${estadisticas.get('costo_por_kg', 0):,.0f}"),
        ]
        
        row += 2
        for label, value in econ:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _crear_hoja_registros(self, ws, registros):
        """Crea la hoja de registros diarios"""
        
        headers = ['Fecha', 'Alimento (kg)', 'Agua (L)', 'Mortalidad', 'Peso (g)', 'Temperatura (°C)', 'Observaciones']
        ws.append(headers)
        
        # Estilo de headers
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Datos
        for reg in registros:
            ws.append([
                reg.get('fecha'),
                reg.get('alimento_kg'),
                reg.get('agua_litros'),
                reg.get('mortalidad', 0),
                reg.get('peso_promedio'),
                reg.get('temperatura_promedio'),
                reg.get('observaciones', '')
            ])
        
        # Ajustar anchos
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['G'].width = 40
    
    def _crear_hoja_costos(self, ws, costos):
        """Crea la hoja de costos"""
        
        headers = ['Fecha', 'Categoría', 'Concepto', 'Monto', 'Observaciones']
        ws.append(headers)
        
        # Estilo de headers
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Datos
        total = 0
        for costo in costos:
            monto = costo.get('monto', 0)
            ws.append([
                costo.get('fecha'),
                costo.get('categoria'),
                costo.get('concepto'),
                monto,
                costo.get('observaciones', '')
            ])
            total += monto
        
        # Total
        row = len(costos) + 2
        ws[f'C{row}'] = 'TOTAL:'
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = total
        ws[f'D{row}'].font = Font(bold=True)
        
        # Ajustar anchos
        for col, width in [('A', 12), ('B', 15), ('C', 30), ('D', 15), ('E', 30)]:
            ws.column_dimensions[col].width = width
    
    def _crear_hoja_sanidad(self, ws, sanidad):
        """Crea la hoja de sanidad"""
        
        headers = ['Fecha', 'Tipo', 'Producto', 'Dosis', 'Edad (días)', 'Observaciones']
        ws.append(headers)
        
        # Estilo de headers
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Datos
        for san in sanidad:
            ws.append([
                san.get('fecha'),
                san.get('tipo'),
                san.get('producto'),
                san.get('dosis'),
                san.get('edad_dias'),
                san.get('observaciones', '')
            ])
        
        # Ajustar anchos
        for col, width in [('A', 12), ('B', 15), ('C', 25), ('D', 15), ('E', 12), ('F', 30)]:
            ws.column_dimensions[col].width = width

class ExportadorCSV:
    """Exporta datos a CSV"""
    
    @staticmethod
    def exportar_registros(registros: list, filename: str):
        """Exporta registros diarios a CSV"""
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow(['Fecha', 'Alimento (kg)', 'Agua (L)', 'Mortalidad', 
                           'Peso (g)', 'Temperatura (°C)', 'Observaciones'])
            
            # Datos
            for reg in registros:
                writer.writerow([
                    reg.get('fecha'),
                    reg.get('alimento_kg'),
                    reg.get('agua_litros'),
                    reg.get('mortalidad', 0),
                    reg.get('peso_promedio'),
                    reg.get('temperatura_promedio'),
                    reg.get('observaciones', '')
                ])
        
        return filename
    
    @staticmethod
    def exportar_costos(costos: list, filename: str):
        """Exporta costos a CSV"""
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow(['Fecha', 'Categoría', 'Concepto', 'Monto', 'Observaciones'])
            
            # Datos
            for costo in costos:
                writer.writerow([
                    costo.get('fecha'),
                    costo.get('categoria'),
                    costo.get('concepto'),
                    costo.get('monto'),
                    costo.get('observaciones', '')
                ])
        
        return filename